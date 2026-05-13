#!/usr/bin/env python3
"""
FTC Communication - Structured Python dicts/lists for PostgreSQL insertion.

Extracts data from 3 Excel files (Apr 28, 29, 30) covering:
  - contd4_records        : Generation capacity pending CONTD-4 approval
  - ftc_records           : FTC/TOC/COD pipeline projects
  - trans_records         : Transmission elements under FTC
  - summary_ftc_pipeline  : Table 2 aggregates (Region x Source x Stage)
  - summary_trans_elements: Table 3 transmission element counts
  - summary_cod_april     : Table 6 April COD by source and region

Usage:
    from ftc_data_dicts import load_all
    data = load_all()
    # data["Apr28"]["contd4_records"]  -> list of dicts ready for INSERT
    # data["Apr28"]["ftc_records"]     -> list of dicts ready for INSERT
    # data["Apr28"]["trans_records"]   -> list of dicts ready for INSERT
    # data["Apr28"]["summary_ftc_pipeline"]  -> aggregated Table-2 dicts
    # data["Apr28"]["summary_trans_elements"] -> aggregated Table-3 dicts
    # data["Apr28"]["summary_cod_april"]      -> aggregated Table-6 dicts
"""

import openpyxl
import datetime

# ─── File paths ────────────────────────────────────────────────────────────────
EXCEL_FILES = {
    "Apr28": "/home/prasad-173/applications/development/ftc-communication/public/data/excel/CONTD and FTC details 280426.xlsx",
    "Apr29": "/home/prasad-173/applications/development/ftc-communication/public/data/excel/CONTD and FTC details 290426.xlsx",
    "Apr30": "/home/prasad-173/applications/development/ftc-communication/public/data/excel/CONTD and FTC details 30.04.xlsx",
}

REGIONAL_SHEETS = ["NR", "WR", "ER", "NER", "SR"]

# ─── Canonical column maps ──────────────────────────────────────────────────────
# Map the messy/variant header strings to canonical DB column names.

CONTD4_COL_MAP = {
    "Sr. No": "sr_no",
    "S.No": "sr_no",
    "S No": "sr_no",
    "Generating Station": "generating_station",
    "Name of Developer": "developer_name",
    "Name of Project": "generating_station",
    "Pooling Station": "pooling_station",
    "Region": "region_col",
    "Generation Type (Wind/Solar/Hybrid/BESS/Coal/Hydro etc)": "generation_type",
    "Capacity(MW)": "capacity_mw",
    "Application Date (dd-mm-yyyy)": "application_date",
    "Application Date": "application_date",
    "Proposed FTC date": "proposed_ftc_date",
    "Capacity(MW) to be completed in Apr'26": "capacity_expected_apr26_mw",
    "Issues if any causing delay/Remark": "issues_remark",
}

FTC_COL_MAP = {
    "Generating Station": "generating_station",
    "Pooling Station": "pooling_station",
    "Plant Type (Wind/Solar/BESS/Hybrid(wind+Solar)/Hybrid(Solar+BESS)/Hybrid(Wind+Solar+BESS).../Coal/Hydro etc)": "plant_type",
    "Region": "region_col",
    "Total Plant Capacity(MW)": "total_capacity_mw",
    "Total Capacity(MW) \n(For which CONTD4 issued)": "contd4_capacity_mw",
    "Capacity(MW) applied for FTC": "applied_ftc_mw",
    "Capacity (MW) applied for FTC": "applied_ftc_mw",
    "Sources Type Applied for FTC(Wind/Solar/BESS/Coal/Hydro/PSP)": "source_type",
    "FTC Completed Capacity(MW)": "ftc_completed_mw",
    "FTC Completed Capacity (MW)": "ftc_completed_mw",
    "FTC approved": "ftc_completed_mw",          # WR uses this label
    "FTC date if completed": "ftc_date",
    "TOC Issued Capacity(MW)": "toc_issued_mw",
    "TOC Issued Capacity (MW)": "toc_issued_mw",
    "TOC issuance date if Completed": "toc_date",
    "COD declared Capacity(MW)": "cod_declared_mw",
    "COD declared Capacity (MW)": "cod_declared_mw",
    "COD Date if Declared": "cod_date",
    "Proposed FTC date if Under process": "proposed_ftc_date",
    "Capacity Under Process for FTC": "pending_ftc_mw",
    "Capacity Under Process for TOC": "pending_toc_mw",
    "Capacity Pending for COD": "pending_cod_mw",
    "Capacity(MW) commisioning expected in Apr'26": "expected_apr26_mw",
    "Issues if any causing delay in FTC/TOC/COD": "issues",
    "Any Otherremark": "other_remark",
    "Any Other remark": "other_remark",
}

TRANS_COL_MAP = {
    "Agency/Owner": "agency_owner",
    "Agency/ Owner": "agency_owner",
    "Name of Line/ICT /GT/ST": "element_name",
    "Type\n(Line/ICT /GT/ST)": "element_type",
    "RE/Non-RE": "re_non_re",
    "Voltage Rating (kV)": "voltage_kv",
    "Capacity (MVA)": "capacity_mva",
    "Line length": "line_length",
    "Date of First Time Energization & Integration": "energization_date",
    "Date of First Time Energization & Integration Approval": "energization_date",
    "Pendig for FTC (Yes/No)": "pending_ftc",
    "Proposed FTC date if Pending": "proposed_ftc_date",
    "Capacity (MVA) to be commissioned in Apr'26": "capacity_apr26_mva",
    "Line Length (ckt km) to be commissioned in Apr'26": "length_apr26_ckt_km",
    "Reason for delay/Any Otherremark": "reason_delay",
}


# ─── Utility helpers ───────────────────────────────────────────────────────────

def _cv(cell):
    """Read and clean a single cell value."""
    v = cell.value
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def _row_vals(row):
    return [_cv(c) for c in row]


def _is_empty(vals):
    return all(v is None for v in vals)


def _clean_headers(vals):
    h = list(vals)
    while h and h[-1] is None:
        h.pop()
    return h


def _serialize(v):
    """Convert datetime → ISO string; leave everything else as-is."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return v


def _detect_section(vals):
    combined = " ".join(str(v) for v in vals if v is not None).lower()
    if "contd-4" in combined:
        return "CONTD4"
    if "transmission elements" in combined:
        return "TRANS"
    if "source-wise segregation" in combined or "hybrid capacity" in combined:
        return "HYBRID"
    if "generation capacity" in combined and "ftc" in combined:
        return "FTC"
    return None


def _map_row(col_map, headers, values):
    """
    Map raw row values to canonical field names using col_map.
    Returns a dict with canonical keys; unmapped columns go into _extra.
    """
    rec = {}
    extra = {}
    for h, v in zip(headers, values):
        if h is None:
            continue
        hstr = str(h).strip()
        canon = col_map.get(hstr)
        sv = _serialize(v)
        if canon:
            if canon not in rec:          # first occurrence wins
                rec[canon] = sv
        else:
            extra[hstr] = sv
    if extra:
        rec["_extra"] = extra
    return rec


# ─── Regional sheet parser ─────────────────────────────────────────────────────

def _parse_regional_sheet(ws):
    """
    Returns a dict keyed by section type ("CONTD4", "FTC", "TRANS", "HYBRID").
    Each value: {"header_row": int, "headers": [...], "rows": [{"row_num":int,"values":[...]}]}
    """
    all_rows = [(i + 1, _row_vals(r)) for i, r in enumerate(ws.iter_rows())]
    sections = {}
    i = 0
    cur_sec = cur_hdr_row = cur_hdrs = None
    cur_data = []

    def _flush():
        nonlocal cur_sec, cur_hdr_row, cur_hdrs, cur_data
        if cur_sec and cur_hdrs is not None:
            sections[cur_sec] = {
                "header_row": cur_hdr_row,
                "headers": cur_hdrs,
                "rows": cur_data,
            }
        cur_sec = cur_hdr_row = cur_hdrs = None
        cur_data = []

    while i < len(all_rows):
        rn, vals = all_rows[i]
        st = _detect_section(vals)
        if st:
            _flush()
            cur_sec = st
            j = i + 1
            while j < len(all_rows) and _is_empty(all_rows[j][1]):
                j += 1
            if j < len(all_rows):
                cur_hdr_row = all_rows[j][0]
                cur_hdrs = _clean_headers(all_rows[j][1])
                i = j + 1
                cur_data = []
            else:
                i += 1
            continue
        if cur_sec and not _is_empty(vals):
            cur_data.append({"row_num": rn, "values": vals})
        i += 1

    _flush()
    return sections


# ─── Build record lists ────────────────────────────────────────────────────────

def _build_contd4(file_label, region, section):
    records = []
    hdrs = section["headers"]
    for r in section["rows"]:
        base = _map_row(CONTD4_COL_MAP, hdrs, r["values"])
        base.update({
            "file_date": file_label,
            "region": region,
            "_sheet_row": r["row_num"],
        })
        records.append(base)
    return records


def _build_ftc(file_label, region, section):
    records = []
    hdrs = section["headers"]
    for r in section["rows"]:
        base = _map_row(FTC_COL_MAP, hdrs, r["values"])
        base.update({
            "file_date": file_label,
            "region": region,
            "_sheet_row": r["row_num"],
        })
        records.append(base)
    return records


def _build_trans(file_label, region, section):
    records = []
    hdrs = section["headers"]
    for r in section["rows"]:
        base = _map_row(TRANS_COL_MAP, hdrs, r["values"])
        base.update({
            "file_date": file_label,
            "region": region,
            "_sheet_row": r["row_num"],
        })
        records.append(base)
    return records


def _build_hybrid(file_label, region, section):
    """Hybrid section varies too much per region; return raw rows."""
    records = []
    hdrs = section["headers"]
    for r in section["rows"]:
        d = {}
        for h, v in zip(hdrs, r["values"]):
            if h is not None:
                d[str(h).strip()] = _serialize(v)
        d["file_date"] = file_label
        d["region"] = region
        d["_sheet_row"] = r["row_num"]
        records.append(d)
    return records


# ─── Summary sheet parsers ─────────────────────────────────────────────────────

def _parse_summary(ws, file_label):
    """
    Parses the Summary sheet and returns three structured tables:
      - ftc_pipeline   : Table 2 rows (region × source × FTC stage) rows 37-85
      - trans_elements : Table 3 rows (region × element type) rows 88-114
      - cod_april      : Table 6 rows (source × region COD) rows 211-220
    """
    all_rows = {i + 1: _row_vals(r) for i, r in enumerate(ws.iter_rows())}

    # ── Table 2: FTC Pipeline (rows 38-85) ─────────────────────────────────
    # Header at row 37:
    # col[0]=Region, col[1]=Source, col[2]=TotalInstalled, col[3]=CONTD4Issued,
    # col[4]=AppliedFTC, col[5]=FTCApproved, col[6]=FTCPending,
    # col[7]=TOCIssued,  col[8]=TOCPending,  col[9]=CODCompleted,
    # col[10]=CODPending, col[11]=ExpApril26

    ftc_pipeline = []
    current_region = None
    for rn in range(38, 86):
        vals = all_rows.get(rn, [])
        if not vals or _is_empty(vals):
            continue
        region_cell = vals[0]
        if region_cell is not None and str(region_cell).strip():
            current_region = str(region_cell).strip()
        source = vals[1] if len(vals) > 1 else None
        if source is None:
            continue

        def fv(idx):
            v = vals[idx] if idx < len(vals) else None
            if isinstance(v, str) and v.strip() in ("#N/A", ""):
                return None
            return _serialize(v)

        ftc_pipeline.append({
            "file_date": file_label,
            "region": current_region,
            "source_type": str(source).strip() if source else None,
            "total_installed_mw": fv(2),
            "contd4_issued_mw": fv(3),
            "applied_ftc_mw": fv(4),
            "ftc_approved_mw": fv(5),
            "ftc_pending_mw": fv(6),
            "toc_issued_mw": fv(7),
            "toc_pending_mw": fv(8),
            "cod_completed_mw": fv(9),
            "cod_pending_mw": fv(10),
            "expected_apr26_mw": fv(11),
            "_sheet_row": rn,
        })

    # ── Table 3: Transmission Elements (rows 91-110) ────────────────────────
    # Header rows 89-90:
    # col[0]=Region, col[1]=ElementType,
    # col[2]=FTCCompleted_cktkmMVA, col[3]=FTCCompleted_totalNo,
    # col[4]=FTCPending_cktkmMVA,   col[5]=FTCPending_totalNo,
    # col[6]=ToBeCommissioned_Apr26_cktkmMVA, col[7]=..._totalNo

    trans_elements = []
    current_region = None
    for rn in range(91, 115):
        vals = all_rows.get(rn, [])
        if not vals or _is_empty(vals):
            continue
        region_cell = vals[0]
        if region_cell is not None and str(region_cell).strip():
            current_region = str(region_cell).strip()
        elem_type = vals[1] if len(vals) > 1 else None
        if elem_type is None:
            continue

        def fv2(idx):
            v = vals[idx] if idx < len(vals) else None
            if isinstance(v, str) and v.strip() in ("#N/A", ""):
                return None
            return _serialize(v)

        trans_elements.append({
            "file_date": file_label,
            "region": current_region,
            "element_type": str(elem_type).strip(),
            "ftc_completed_cktkm_mva": fv2(2),
            "ftc_completed_count": fv2(3),
            "ftc_pending_cktkm_mva": fv2(4),
            "ftc_pending_count": fv2(5),
            "commissioning_apr26_cktkm_mva": fv2(6),
            "commissioning_apr26_count": fv2(7),
            "_sheet_row": rn,
        })

    # ── Table 6: COD Declared in April (rows 211-220) ───────────────────────
    # Header at row 212: Source | NR | WR | SR | ER | NER | All India
    REGIONS_ORDER = ["NR", "WR", "SR", "ER", "NER", "All India"]
    cod_april = []
    for rn in range(213, 221):
        vals = all_rows.get(rn, [])
        if not vals or _is_empty(vals):
            continue
        source = vals[0]
        if source is None:
            continue
        row_rec = {"file_date": file_label, "source_type": str(source).strip(), "_sheet_row": rn}
        for col_idx, region in enumerate(REGIONS_ORDER, start=1):
            v = vals[col_idx] if col_idx < len(vals) else None
            # Some Hybrid cells have text like "180 - BESS\n211.4 - Solar\n40.95 - Wind"
            row_rec[region.lower().replace(" ", "_") + "_mw"] = _serialize(v)
        cod_april.append(row_rec)

    return {
        "ftc_pipeline": ftc_pipeline,
        "trans_elements": trans_elements,
        "cod_april": cod_april,
    }


# ─── Main loader ───────────────────────────────────────────────────────────────

def load_file(file_label: str, filepath: str) -> dict:
    """
    Load one Excel file and return structured data:
    {
      "file_date": str,
      "sheet_names": [...],
      "contd4_records": [...],
      "ftc_records": [...],
      "trans_records": [...],
      "hybrid_records": [...],
      "summary_ftc_pipeline": [...],
      "summary_trans_elements": [...],
      "summary_cod_april": [...],
    }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    contd4_records = []
    ftc_records = []
    trans_records = []
    hybrid_records = []

    for sheet_name in REGIONAL_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        sections = _parse_regional_sheet(ws)

        if "CONTD4" in sections:
            contd4_records.extend(_build_contd4(file_label, sheet_name, sections["CONTD4"]))
        if "FTC" in sections:
            ftc_records.extend(_build_ftc(file_label, sheet_name, sections["FTC"]))
        if "TRANS" in sections:
            trans_records.extend(_build_trans(file_label, sheet_name, sections["TRANS"]))
        if "HYBRID" in sections:
            hybrid_records.extend(_build_hybrid(file_label, sheet_name, sections["HYBRID"]))

    summary = {"ftc_pipeline": [], "trans_elements": [], "cod_april": []}
    if "Summary" in wb.sheetnames:
        summary = _parse_summary(wb["Summary"], file_label)

    wb.close()

    return {
        "file_date": file_label,
        "sheet_names": wb.sheetnames if hasattr(wb, "sheetnames") else [],
        "contd4_records": contd4_records,
        "ftc_records": ftc_records,
        "trans_records": trans_records,
        "hybrid_records": hybrid_records,
        "summary_ftc_pipeline": summary["ftc_pipeline"],
        "summary_trans_elements": summary["trans_elements"],
        "summary_cod_april": summary["cod_april"],
    }


def load_all() -> dict:
    """
    Load all three files. Returns:
    {
      "Apr28": { ... },
      "Apr29": { ... },
      "Apr30": { ... },
    }
    """
    result = {}
    for label, fpath in EXCEL_FILES.items():
        result[label] = load_file(label, fpath)
    return result


# ─── Suggested PostgreSQL schema ───────────────────────────────────────────────

POSTGRES_DDL = """
-- Run once to create tables

CREATE TABLE IF NOT EXISTS ftc_contd4 (
    id                      SERIAL PRIMARY KEY,
    file_date               VARCHAR(10),          -- 'Apr28', 'Apr29', 'Apr30'
    region                  VARCHAR(10),          -- 'NR','WR','ER','NER','SR'
    sr_no                   NUMERIC,
    developer_name          TEXT,
    generating_station      TEXT,
    pooling_station         TEXT,
    region_col              VARCHAR(10),
    generation_type         TEXT,
    capacity_mw             NUMERIC,
    application_date        TEXT,
    proposed_ftc_date       TEXT,
    capacity_expected_apr26_mw NUMERIC,
    issues_remark           TEXT,
    _sheet_row              INTEGER,
    _extra                  JSONB
);

CREATE TABLE IF NOT EXISTS ftc_pipeline (
    id                      SERIAL PRIMARY KEY,
    file_date               VARCHAR(10),
    region                  VARCHAR(10),
    generating_station      TEXT,
    pooling_station         TEXT,
    plant_type              TEXT,
    region_col              VARCHAR(10),
    total_capacity_mw       NUMERIC,
    contd4_capacity_mw      NUMERIC,
    applied_ftc_mw          NUMERIC,
    source_type             TEXT,
    ftc_completed_mw        NUMERIC,
    ftc_date                TEXT,
    toc_issued_mw           NUMERIC,
    toc_date                TEXT,
    cod_declared_mw         NUMERIC,
    cod_date                TEXT,
    proposed_ftc_date       TEXT,
    pending_ftc_mw          NUMERIC,
    pending_toc_mw          NUMERIC,
    pending_cod_mw          NUMERIC,
    expected_apr26_mw       NUMERIC,
    issues                  TEXT,
    other_remark            TEXT,
    _sheet_row              INTEGER,
    _extra                  JSONB
);

CREATE TABLE IF NOT EXISTS ftc_trans_elements (
    id                      SERIAL PRIMARY KEY,
    file_date               VARCHAR(10),
    region                  VARCHAR(10),
    agency_owner            TEXT,
    element_name            TEXT,
    element_type            TEXT,
    re_non_re               VARCHAR(20),
    voltage_kv              TEXT,
    capacity_mva            TEXT,
    line_length             TEXT,
    energization_date       TEXT,
    pending_ftc             TEXT,
    proposed_ftc_date       TEXT,
    capacity_apr26_mva      TEXT,
    length_apr26_ckt_km     TEXT,
    reason_delay            TEXT,
    _sheet_row              INTEGER,
    _extra                  JSONB
);

CREATE TABLE IF NOT EXISTS ftc_summary_pipeline (
    id                      SERIAL PRIMARY KEY,
    file_date               VARCHAR(10),
    region                  VARCHAR(20),
    source_type             VARCHAR(50),
    total_installed_mw      NUMERIC,
    contd4_issued_mw        NUMERIC,
    applied_ftc_mw          NUMERIC,
    ftc_approved_mw         NUMERIC,
    ftc_pending_mw          NUMERIC,
    toc_issued_mw           NUMERIC,
    toc_pending_mw          NUMERIC,
    cod_completed_mw        NUMERIC,
    cod_pending_mw          NUMERIC,
    expected_apr26_mw       NUMERIC,
    _sheet_row              INTEGER
);

CREATE TABLE IF NOT EXISTS ftc_summary_trans (
    id                          SERIAL PRIMARY KEY,
    file_date                   VARCHAR(10),
    region                      VARCHAR(20),
    element_type                VARCHAR(100),
    ftc_completed_cktkm_mva     NUMERIC,
    ftc_completed_count         NUMERIC,
    ftc_pending_cktkm_mva       NUMERIC,
    ftc_pending_count           NUMERIC,
    commissioning_apr26_cktkm_mva NUMERIC,
    commissioning_apr26_count   NUMERIC,
    _sheet_row                  INTEGER
);

CREATE TABLE IF NOT EXISTS ftc_summary_cod_april (
    id              SERIAL PRIMARY KEY,
    file_date       VARCHAR(10),
    source_type     VARCHAR(50),
    nr_mw           TEXT,
    wr_mw           TEXT,
    sr_mw           TEXT,
    er_mw           TEXT,
    ner_mw          TEXT,
    all_india_mw    TEXT,
    _sheet_row      INTEGER
);
"""

# ─── Sample INSERT helpers ─────────────────────────────────────────────────────

def insert_contd4(conn, records):
    """
    Insert contd4_records into ftc_contd4.
    conn: psycopg2 connection
    """
    import json
    cur = conn.cursor()
    sql = """
        INSERT INTO ftc_contd4 (
            file_date, region, sr_no, developer_name, generating_station,
            pooling_station, region_col, generation_type, capacity_mw,
            application_date, proposed_ftc_date, capacity_expected_apr26_mw,
            issues_remark, _sheet_row, _extra
        ) VALUES (
            %(file_date)s, %(region)s, %(sr_no)s, %(developer_name)s, %(generating_station)s,
            %(pooling_station)s, %(region_col)s, %(generation_type)s, %(capacity_mw)s,
            %(application_date)s, %(proposed_ftc_date)s, %(capacity_expected_apr26_mw)s,
            %(issues_remark)s, %(_sheet_row)s, %(extra_json)s
        )
    """
    for rec in records:
        row = dict(rec)
        row["developer_name"] = row.get("developer_name")
        row["extra_json"] = json.dumps(row.pop("_extra", {}))
        cur.execute(sql, row)
    conn.commit()
    cur.close()


def insert_ftc(conn, records):
    import json
    cur = conn.cursor()
    sql = """
        INSERT INTO ftc_pipeline (
            file_date, region, generating_station, pooling_station, plant_type,
            region_col, total_capacity_mw, contd4_capacity_mw, applied_ftc_mw,
            source_type, ftc_completed_mw, ftc_date, toc_issued_mw, toc_date,
            cod_declared_mw, cod_date, proposed_ftc_date, pending_ftc_mw,
            pending_toc_mw, pending_cod_mw, expected_apr26_mw, issues, other_remark,
            _sheet_row, _extra
        ) VALUES (
            %(file_date)s, %(region)s, %(generating_station)s, %(pooling_station)s,
            %(plant_type)s, %(region_col)s, %(total_capacity_mw)s, %(contd4_capacity_mw)s,
            %(applied_ftc_mw)s, %(source_type)s, %(ftc_completed_mw)s, %(ftc_date)s,
            %(toc_issued_mw)s, %(toc_date)s, %(cod_declared_mw)s, %(cod_date)s,
            %(proposed_ftc_date)s, %(pending_ftc_mw)s, %(pending_toc_mw)s,
            %(pending_cod_mw)s, %(expected_apr26_mw)s, %(issues)s, %(other_remark)s,
            %(_sheet_row)s, %(extra_json)s
        )
    """
    for rec in records:
        row = dict(rec)
        row["extra_json"] = json.dumps(row.pop("_extra", {}))
        cur.execute(sql, row)
    conn.commit()
    cur.close()


def insert_trans(conn, records):
    import json
    cur = conn.cursor()
    sql = """
        INSERT INTO ftc_trans_elements (
            file_date, region, agency_owner, element_name, element_type,
            re_non_re, voltage_kv, capacity_mva, line_length, energization_date,
            pending_ftc, proposed_ftc_date, capacity_apr26_mva, length_apr26_ckt_km,
            reason_delay, _sheet_row, _extra
        ) VALUES (
            %(file_date)s, %(region)s, %(agency_owner)s, %(element_name)s, %(element_type)s,
            %(re_non_re)s, %(voltage_kv)s, %(capacity_mva)s, %(line_length)s,
            %(energization_date)s, %(pending_ftc)s, %(proposed_ftc_date)s,
            %(capacity_apr26_mva)s, %(length_apr26_ckt_km)s, %(reason_delay)s,
            %(_sheet_row)s, %(extra_json)s
        )
    """
    for rec in records:
        row = dict(rec)
        row["extra_json"] = json.dumps(row.pop("_extra", {}))
        cur.execute(sql, row)
    conn.commit()
    cur.close()


def insert_summary_pipeline(conn, records):
    cur = conn.cursor()
    sql = """
        INSERT INTO ftc_summary_pipeline (
            file_date, region, source_type, total_installed_mw, contd4_issued_mw,
            applied_ftc_mw, ftc_approved_mw, ftc_pending_mw, toc_issued_mw,
            toc_pending_mw, cod_completed_mw, cod_pending_mw, expected_apr26_mw,
            _sheet_row
        ) VALUES (
            %(file_date)s, %(region)s, %(source_type)s, %(total_installed_mw)s,
            %(contd4_issued_mw)s, %(applied_ftc_mw)s, %(ftc_approved_mw)s,
            %(ftc_pending_mw)s, %(toc_issued_mw)s, %(toc_pending_mw)s,
            %(cod_completed_mw)s, %(cod_pending_mw)s, %(expected_apr26_mw)s,
            %(_sheet_row)s
        )
    """
    cur.executemany(sql, records)
    conn.commit()
    cur.close()


def insert_summary_trans(conn, records):
    cur = conn.cursor()
    sql = """
        INSERT INTO ftc_summary_trans (
            file_date, region, element_type, ftc_completed_cktkm_mva,
            ftc_completed_count, ftc_pending_cktkm_mva, ftc_pending_count,
            commissioning_apr26_cktkm_mva, commissioning_apr26_count, _sheet_row
        ) VALUES (
            %(file_date)s, %(region)s, %(element_type)s, %(ftc_completed_cktkm_mva)s,
            %(ftc_completed_count)s, %(ftc_pending_cktkm_mva)s, %(ftc_pending_count)s,
            %(commissioning_apr26_cktkm_mva)s, %(commissioning_apr26_count)s, %(_sheet_row)s
        )
    """
    cur.executemany(sql, records)
    conn.commit()
    cur.close()


def insert_summary_cod(conn, records):
    cur = conn.cursor()
    sql = """
        INSERT INTO ftc_summary_cod_april (
            file_date, source_type, nr_mw, wr_mw, sr_mw, er_mw, ner_mw, all_india_mw, _sheet_row
        ) VALUES (
            %(file_date)s, %(source_type)s, %(nr_mw)s, %(wr_mw)s, %(sr_mw)s,
            %(er_mw)s, %(ner_mw)s, %(all_india_mw)s, %(_sheet_row)s
        )
    """
    cur.executemany(sql, records)
    conn.commit()
    cur.close()


# ─── Quick-verify / demo ───────────────────────────────────────────────────────

if __name__ == "__main__":
    import json

    print("Loading all 3 files...")
    data = load_all()

    for label in ["Apr28", "Apr29", "Apr30"]:
        d = data[label]
        print(f"\n{'='*60}")
        print(f"FILE: {label}")
        print(f"{'='*60}")
        print(f"  contd4_records        : {len(d['contd4_records'])} rows")
        print(f"  ftc_records           : {len(d['ftc_records'])} rows")
        print(f"  trans_records         : {len(d['trans_records'])} rows")
        print(f"  hybrid_records        : {len(d['hybrid_records'])} rows")
        print(f"  summary_ftc_pipeline  : {len(d['summary_ftc_pipeline'])} rows")
        print(f"  summary_trans_elements: {len(d['summary_trans_elements'])} rows")
        print(f"  summary_cod_april     : {len(d['summary_cod_april'])} rows")

    # Print sample records from Apr28
    print("\n\n--- SAMPLE contd4_records[0] (Apr28 NR) ---")
    print(json.dumps(data["Apr28"]["contd4_records"][0], indent=2, default=str))

    print("\n--- SAMPLE ftc_records[0] (Apr28 NR) ---")
    print(json.dumps(data["Apr28"]["ftc_records"][0], indent=2, default=str))

    print("\n--- SAMPLE trans_records[0] (Apr28 NR) ---")
    print(json.dumps(data["Apr28"]["trans_records"][0], indent=2, default=str))

    print("\n--- SAMPLE summary_ftc_pipeline rows (NR + WR + SR + ER + NER Total rows) ---")
    totals = [r for r in data["Apr28"]["summary_ftc_pipeline"] if r["source_type"] == "Total"]
    for t in totals:
        print(json.dumps(t, indent=2, default=str))

    print("\n--- SAMPLE summary_cod_april (Apr28) ---")
    for r in data["Apr28"]["summary_cod_april"]:
        print(json.dumps(r, indent=2, default=str))

    print("\n--- SAMPLE summary_trans_elements (Apr28, first 6 rows) ---")
    for r in data["Apr28"]["summary_trans_elements"][:6]:
        print(json.dumps(r, indent=2, default=str))

    print("\n\nPostgreSQL DDL available in ftc_data_dicts.POSTGRES_DDL")
    print("Insert helpers: insert_contd4(), insert_ftc(), insert_trans(),")
    print("                insert_summary_pipeline(), insert_summary_trans(), insert_summary_cod()")
