"""
Tokenise the multi-date FTC / TOC / COD cells (columns K, M, O of per-region
Section 2 sheets) into individual (capacity_mw, event_date) tuples and dump them
to JSON. By default this is a *dry-run*: it prints the per-project parse and
writes events_parsed.json. To actually insert the events, run scripts/seed_events_to_db.py
on the resulting JSON.

Usage:
    python3 scripts/parse_events_from_excel.py                     # latest file, dry-run
    python3 scripts/parse_events_from_excel.py --file <name>       # specific file
    python3 scripts/parse_events_from_excel.py --raw               # also dump raw cell strings

Output formats encountered in cells (all handled):
  - "66.66 FTC completed on 26-02-2026"
  - "33.33 MW-17 MAR 26"
  - "150MW (30.03.2026)"
  - "60MW-23-03-2026" / "60MW  dated 17.03.2026"
  - "50: 25-01-2026" (MW: date)
  - "09-01-2026 (156.88 MW)" (date first)
  - "Solar:16.03.2026 (185MW)" / "BESS: 17.03.2026 (90MW)" — labelled
  - "Unit-2: 02.12.2025" — no MW, skipped (cannot attribute quantum)
"""

import json
import re
import sys
import argparse
from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook

DATA_DIR = Path(__file__).parent.parent / "public/data/excel"
DEFAULT_FILE = "CONTD and FTC details 130526.xlsx"

# ── Regex helpers ────────────────────────────────────────────────────────────

# Date forms:  dd-mm-yyyy  dd.mm.yyyy  dd/mm/yyyy  with optional 2-digit year
# Or:           dd MMM yyyy  / dd MMM yy
MONTH_NAMES = {
    'jan': 1, 'january': 1,
    'feb': 2, 'february': 2,
    'mar': 3, 'march': 3,
    'apr': 4, 'april': 4,
    'may': 5,
    'jun': 6, 'june': 6,
    'jul': 7, 'july': 7,
    'aug': 8, 'august': 8,
    'sep': 9, 'sept': 9, 'september': 9,
    'oct': 10, 'october': 10,
    'nov': 11, 'november': 11,
    'dec': 12, 'december': 12,
}

DATE_RX = re.compile(
    r'(?P<dmY>\b(\d{1,2})\s*[-./\s]\s*(\d{1,2})\s*[-./\s]\s*(\d{2,4})\b)'
    r'|'
    r'(?P<dmName>\b(\d{1,2})\s+([A-Za-z]{3,9})\s+(\d{2,4})\b)'
)

MW_RX = re.compile(r'(\d+(?:\.\d+)?)\s*MW', re.IGNORECASE)
# Bare numeric (used when format is "50: 25-01-2026")
LEADING_NUM_RX = re.compile(r'^\s*(\d+(?:\.\d+)?)\s*[:\-]')

def _norm_year(y):
    yi = int(y)
    if yi < 50:
        return 2000 + yi
    if yi < 100:
        return 1900 + yi
    return yi

def parse_date(m):
    """Convert a DATE_RX match into datetime.date, or None."""
    if m.group('dmY'):
        d, mo, y = int(m.group(2)), int(m.group(3)), _norm_year(m.group(4))
    elif m.group('dmName'):
        d = int(m.group(6))
        mo = MONTH_NAMES.get(m.group(7).lower())
        if not mo: return None
        y = _norm_year(m.group(8))
    else:
        return None
    # If parsed month > 12, swap (some entries are m-d-y by mistake)
    if mo > 12 and d <= 12:
        d, mo = mo, d
    if not (1 <= mo <= 12 and 1 <= d <= 31 and 2020 <= y <= 2035):
        return None
    try:
        return date(y, mo, d)
    except ValueError:
        return None

def tokenize_cell(raw):
    """Split a multi-entry cell into individual fragments.
    Newlines are firm separators; multiple-spaces also split when fragments
    look fully terminated (i.e. contain both MW and a date)."""
    if raw is None:
        return []
    s = str(raw).replace('\r', '').strip()
    if not s:
        return []
    # First split on newlines.
    frags = [f.strip() for f in s.split('\n') if f.strip()]
    out = []
    for f in frags:
        # Heuristic: if the fragment contains >1 MW occurrence AND >1 date,
        # it's multiple entries concatenated by whitespace — split on "  " (>=2 spaces).
        if len(MW_RX.findall(f)) > 1 and len(DATE_RX.findall(f)) > 1:
            # Greedy split: pair MW with the nearest following date in the string.
            # We split on sequences of >=2 spaces or tab to break stuck-together entries.
            sub = re.split(r'\s{2,}|\t+', f)
            for s2 in sub:
                if s2.strip():
                    out.append(s2.strip())
        else:
            out.append(f)
    return out

def parse_fragment(frag):
    """Extract (mw, date) from one fragment. Returns None if either is missing."""
    if not frag:
        return None
    # Skip obvious header/label-only fragments
    fl = frag.strip().lower()
    if fl in ('solar', 'wind', 'bess', 'hybrid'):
        return None

    # Try "Number: date" leading form (e.g. "50: 25-01-2026")
    mw_val = None
    lead = LEADING_NUM_RX.match(frag)
    if lead and 'MW' not in frag.upper():
        mw_val = float(lead.group(1))

    # MW pattern (e.g. "33.33 MW", "133.33MW")
    if mw_val is None:
        mw_m = MW_RX.search(frag)
        if mw_m:
            mw_val = float(mw_m.group(1))

    # Fallback: leading bare number followed by a milestone label
    # ("66.66 FTC completed on 26-02-2026") — value implicit as MW.
    if mw_val is None:
        bare = re.match(r'^\s*(\d+(?:\.\d+)?)\s+(?:FTC|TOC|COD)\b', frag, re.IGNORECASE)
        if bare:
            mw_val = float(bare.group(1))

    # Date pattern (first hit wins)
    date_m = DATE_RX.search(frag)
    d_val = parse_date(date_m) if date_m else None

    if mw_val is None or d_val is None:
        return None
    return (mw_val, d_val)

def parse_events_cell(raw):
    """Parse a full cell into a list of {mw, date, raw_fragment} dicts."""
    events = []
    skipped = []
    for frag in tokenize_cell(raw):
        result = parse_fragment(frag)
        if result:
            mw, d = result
            events.append({"mw": mw, "date": d.isoformat(), "fragment": frag})
        else:
            skipped.append(frag)
    return events, skipped


# ── Section 2 row extractor ──────────────────────────────────────────────────

def find_section2_header(ws):
    """Return row number of the column header line, or None."""
    for r in range(1, 100):
        v = ws.cell(row=r, column=4).value
        if v and "Plant Type" in str(v):
            return r
    return None

def parse_region_sheet(ws, region_code):
    header_row = find_section2_header(ws)
    if header_row is None:
        return []

    rows = []
    for r in range(header_row + 1, header_row + 60):
        name = ws.cell(row=r, column=2).value
        if not name: continue
        name_s = str(name).strip()
        if any(stop in name_s for stop in ('Transmission', 'Hybrid Capacity', 'Sr. No', 'Source-wise')):
            break
        plant_type = ws.cell(row=r, column=4).value
        if not plant_type:
            continue
        ftc_evs, ftc_skip = parse_events_cell(ws.cell(row=r, column=11).value)
        toc_evs, toc_skip = parse_events_cell(ws.cell(row=r, column=13).value)
        cod_evs, cod_skip = parse_events_cell(ws.cell(row=r, column=15).value)
        rows.append({
            "region":      region_code,
            "row":         r,
            "name":        name_s,
            "plant_type":  str(plant_type).strip(),
            "source_types": (str(ws.cell(row=r, column=9).value).strip() if ws.cell(row=r, column=9).value else None),
            "ftc_total":   _num(ws.cell(row=r, column=10).value),
            "toc_total":   _num(ws.cell(row=r, column=12).value),
            "cod_total":   _num(ws.cell(row=r, column=14).value),
            "ftc_events":  ftc_evs,
            "toc_events":  toc_evs,
            "cod_events":  cod_evs,
            "ftc_skipped": ftc_skip,
            "toc_skipped": toc_skip,
            "cod_skipped": cod_skip,
            "ftc_raw":     str(ws.cell(row=r, column=11).value) if ws.cell(row=r, column=11).value else None,
            "toc_raw":     str(ws.cell(row=r, column=13).value) if ws.cell(row=r, column=13).value else None,
            "cod_raw":     str(ws.cell(row=r, column=15).value) if ws.cell(row=r, column=15).value else None,
        })
    return rows

def _num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).strip())
    except ValueError: return None


# ── Main / reporting ─────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser()
    p.add_argument('--file', default=DEFAULT_FILE)
    p.add_argument('--raw',  action='store_true', help='also dump raw cell strings to console')
    args = p.parse_args()

    wb = load_workbook(DATA_DIR / args.file, data_only=True)
    all_rows = []
    totals = {'ftc': 0, 'toc': 0, 'cod': 0}
    sums   = {'ftc': 0.0, 'toc': 0.0, 'cod': 0.0}
    skipped_count = 0

    for region in ('NR', 'WR', 'SR', 'ER', 'NER'):
        ws = wb[region]
        rows = parse_region_sheet(ws, region)
        all_rows.extend(rows)
        print(f"\n══ {region} ══  ({len(rows)} projects)")
        for row in rows:
            n = row['name'][:48]
            ft = f"FTC:{len(row['ftc_events']):>2}"
            tt = f"TOC:{len(row['toc_events']):>2}"
            ct = f"COD:{len(row['cod_events']):>2}"
            ft_sum = sum(e['mw'] for e in row['ftc_events'])
            tc_sum = sum(e['mw'] for e in row['toc_events'])
            cd_sum = sum(e['mw'] for e in row['cod_events'])
            # Compare to Excel totals (col 10/12/14)
            mismatch = []
            for k, parsed, total in (
                ('ftc', ft_sum, row['ftc_total']),
                ('toc', tc_sum, row['toc_total']),
                ('cod', cd_sum, row['cod_total']),
            ):
                if total is not None and parsed > 0 and abs(parsed - total) > 0.5:
                    mismatch.append(f"{k.upper()} sum={parsed:.1f} vs total={total:.1f}")
            warn = '  ⚠ ' + '; '.join(mismatch) if mismatch else ''
            print(f"  {n:48}  {ft}  {tt}  {ct}  Σ ftc={ft_sum:>7.2f} toc={tc_sum:>7.2f} cod={cd_sum:>7.2f}{warn}")
            for k in ('ftc', 'toc', 'cod'):
                totals[k] += len(row[f'{k}_events'])
                sums[k]   += sum(e['mw'] for e in row[f'{k}_events'])
                skipped_count += len(row[f'{k}_skipped'])
            if args.raw:
                for k in ('ftc','toc','cod'):
                    if row[f'{k}_raw']:
                        print(f"    raw {k.upper()}: {row[f'{k}_raw'][:140]!r}")
                for k in ('ftc','toc','cod'):
                    if row[f'{k}_skipped']:
                        print(f"    SKIPPED {k.upper()}: {row[f'{k}_skipped']}")

    print("\n══ Summary ══")
    print(f"  Total projects parsed: {len(all_rows)}")
    print(f"  Events parsed:    FTC={totals['ftc']:>4}  TOC={totals['toc']:>4}  COD={totals['cod']:>4}")
    print(f"  Sum of MW:        FTC={sums['ftc']:>7.1f}  TOC={sums['toc']:>7.1f}  COD={sums['cod']:>7.1f}")
    print(f"  Unparseable fragments (skipped): {skipped_count}")

    out = Path(__file__).parent / 'events_parsed.json'
    out.write_text(json.dumps({"source_file": args.file, "rows": all_rows}, indent=2, default=str))
    print(f"\n✓ Dumped parse results to {out.relative_to(Path.cwd()) if out.is_relative_to(Path.cwd()) else out}")
    print("  Next step: review events_parsed.json, then run seed_events_to_db.py --commit")

if __name__ == '__main__':
    main()
