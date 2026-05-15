"""
Parse Section 4 ("Source-wise Segregation of Hybrid Capacity") from each
per-region sheet of the May-13 Excel and update the DB:
  - One CommissioningPhase per source component (Solar, Wind, BESS, PSP)
    for every hybrid project, with the component's own Applied/FTC/TOC/COD MW
  - Project's windCapacityMw / solarCapacityMw / bessCapacityMw

This is what the Excel Summary Hybrid-Breakdown table aggregates from.
Currently our DB stores ONE phase per hybrid project with the totals
collapsed onto a single source type, so the Hybrid Breakdown tab in
the portal shows only one row per project.

Run: python3 scripts/sync_hybrid_components.py [--dry-run]
"""

import sys, re
from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook
import psycopg2, psycopg2.extras

DATA  = Path(__file__).parent.parent / "public/data/excel"
FILE  = "CONTD and FTC details 130526.xlsx"
DB    = "postgresql://postgres:S0perg%4026@10.5.133.55:5432/ftc_communication"
DRY   = "--dry-run" in sys.argv

# Section 4 columns (uniform across regions):
#   C2  Generating Station
#   C3  Pooling Station
#   C4  Plant Type (hybrid label)
#   C5  Region
#   C6  Total Plant Capacity (per component)
#   C7  Total Capacity for which CONTD-4 issued
#   C8  Capacity applied for FTC
#   C9  Source Type (Wind / Solar / BESS / PSP)
#   C10 FTC Completed
#   C11 FTC date
#   C12 TOC Issued
#   C13 TOC date
#   C14 COD declared
#   C15 COD date
#   C17 Under FTC
#   C18 Under TOC
#   C19 COD Pending
#   C20 Expected capacity

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    if not s or s in ('-', 'NA', 'N/A', '0', '0.0'): return 0.0 if s in ('0','0.0') else None
    try: return float(s)
    except ValueError: return None

def text(v):
    if v is None: return None
    s = str(v).strip()
    return s or None

def to_date(v):
    if v is None: return None
    if isinstance(v, (datetime, date)): return v
    s = str(v).strip()
    m = re.match(r'(\d{1,2})[-./](\d{1,2})[-./](\d{4})', s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try: return datetime(y, mo, d)
        except: return None
    return None

def normalise_src(s):
    if not s: return None
    u = s.upper().strip()
    if 'PSP' in u or 'PUMP' in u: return 'PSP'
    if 'WIND' in u:                return 'WIND'
    if 'SOLAR' in u:               return 'SOLAR'
    if 'BESS' in u or 'BATTERY' in u: return 'BESS'
    if 'HYDRO' in u:               return 'HYDRO'
    return None

# ── Parse Section 4 ──────────────────────────────────────────────────────────

def find_section4_header(ws):
    """Find the column-header row for Section 4 (Source-wise Hybrid).

    Strategy:
      1. If a "Source wise Segregation" banner is present, the next row is it.
      2. Otherwise: there are TWO header rows in each per-region sheet that
         contain `Generating Station` + `Plant Type` — Section 2 (FTC) and
         Section 4 (Source-wise Hybrid). Section 4 is always the SECOND one.
    """
    # 1) explicit banner
    for r in range(1, ws.max_row + 1):
        v3 = ws.cell(row=r, column=3).value
        if v3 and 'Source wise Segregation' in str(v3):
            # Header row sits 1-3 rows below the banner
            for delta in range(1, 4):
                v2 = ws.cell(row=r + delta, column=2).value
                if v2 and 'Generating Station' in str(v2):
                    return r + delta
            return r + 1

    # 2) second occurrence of the column header pattern
    hits = []
    for r in range(1, ws.max_row + 1):
        v2 = ws.cell(row=r, column=2).value
        v4 = ws.cell(row=r, column=4).value
        v9 = ws.cell(row=r, column=9).value
        if v2 and 'Generating Station' in str(v2) \
           and v4 and 'Plant Type' in str(v4) \
           and v9 and 'Source' in str(v9):
            hits.append(r)
    return hits[1] if len(hits) >= 2 else None

def parse_section4(ws, region):
    header = find_section4_header(ws)
    if header is None: return []
    rows = []
    current_station = None
    current_pooling = None
    current_plant   = None
    for r in range(header + 1, header + 60):
        name    = text(ws.cell(row=r, column=2).value)
        pooling = text(ws.cell(row=r, column=3).value)
        plant   = text(ws.cell(row=r, column=4).value)
        # If we hit a new section banner, stop.
        all_txt = ' '.join(str(ws.cell(row=r, column=c).value or '') for c in range(1, 6))
        if 'old source' in all_txt.lower(): break

        # A new project row always has C2 (station name) populated.
        if name:
            current_station = name
            current_pooling = pooling
            current_plant   = plant
        # Continuation rows have C2 empty but C4 (plant type) and C9 (source) populated.
        elif plant:
            current_plant = plant

        src = normalise_src(text(ws.cell(row=r, column=9).value))
        if not src or not current_station:
            continue

        rows.append({
            "region":      region,
            "station":     current_station,
            "pooling":     current_pooling,
            "plant_type":  current_plant,
            "source":      src,
            "comp_total":  num(ws.cell(row=r, column=6).value),
            "applied":     num(ws.cell(row=r, column=8).value),
            "ftc":         num(ws.cell(row=r, column=10).value),
            "ftc_date":    to_date(ws.cell(row=r, column=11).value),
            "toc":         num(ws.cell(row=r, column=12).value),
            "toc_date":    to_date(ws.cell(row=r, column=13).value),
            "cod":         num(ws.cell(row=r, column=14).value),
            "cod_date":    to_date(ws.cell(row=r, column=15).value),
            "under_ftc":   num(ws.cell(row=r, column=17).value),
            "under_toc":   num(ws.cell(row=r, column=18).value),
            "cod_pending": num(ws.cell(row=r, column=19).value),
            "expected":    num(ws.cell(row=r, column=20).value),
        })
    return rows

# ── DB update ────────────────────────────────────────────────────────────────

def cuid_like():
    import time, random, string
    chars = string.ascii_lowercase + string.digits
    ts = format(int(time.time() * 1000), 'x')
    return f"cmp{ts}{''.join(random.choices(chars, k=18))}"[:25]


def main():
    print(f"\n{'='*78}\n  Sync hybrid components from Section 4 {'(DRY RUN)' if DRY else ''}\n{'='*78}")
    wb = load_workbook(DATA / FILE, data_only=True)

    all_rows = []
    for region in ['NR', 'WR', 'SR', 'ER', 'NER']:
        if region not in wb.sheetnames:
            continue
        rows = parse_section4(wb[region], region)
        print(f"\n{region}: {len(rows)} hybrid component rows in Section 4")
        for r in rows:
            print(f"  {r['station'][:40]:40}  src={r['source']:5}  total={r['comp_total'] or 0:>7.2f}  "
                  f"app={r['applied'] or 0:>7.2f}  ftc={r['ftc'] or 0:>6.2f}  toc={r['toc'] or 0:>6.2f}  cod={r['cod'] or 0:>6.2f}")
        all_rows.extend(rows)

    print(f"\nTotal Section-4 component rows parsed: {len(all_rows)}")

    # Group rows by (region, station) so each project gets one transaction
    by_project = {}
    for r in all_rows:
        key = (r['region'], r['station'])
        by_project.setdefault(key, []).append(r)

    conn = psycopg2.connect(DB)
    conn.autocommit = False
    cur  = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    updated = 0
    not_found = []

    # Plant-type-code resolver — used when inserting a new hybrid project.
    def hybrid_plant_code(comp_rows):
        srcs = set(r['source'] for r in comp_rows)
        if srcs == {'SOLAR', 'WIND', 'BESS'}: return 'HYBRID_WSB'
        if srcs == {'SOLAR', 'WIND'}:         return 'HYBRID_WS'
        if srcs == {'SOLAR', 'BESS'}:         return 'HYBRID_SB'
        if srcs == {'WIND',  'BESS'}:         return 'HYBRID_WB'
        if srcs == {'WIND',  'PSP'}:          return 'HYBRID_WP'
        if srcs == {'SOLAR', 'PSP'}:          return 'HYBRID_SP'
        if srcs == {'HYDRO', 'PSP'}:          return 'HYBRID_HP'
        return 'HYBRID_WS'  # safe default

    # Look up an admin user once (needed for new project createdById)
    cur.execute("SELECT id FROM users LIMIT 1")
    admin_user = cur.fetchone()['id']

    inserted = 0
    for (region, station), comp_rows in by_project.items():
        first_token = station[:30]
        cur.execute("""
            SELECT gp.id, gp.name, gp."plantTypeId", pt.code AS pt_code, pt."isHybrid"
            FROM generation_projects gp
            JOIN grid_regions gr ON gp."regionId" = gr.id
            JOIN plant_types pt ON gp."plantTypeId" = pt.id
            LEFT JOIN contd4_applications ca ON ca."projectId" = gp.id
            WHERE gr.code = %s AND gp.name ILIKE %s
              AND pt."isHybrid" = TRUE
        """, (region, f"%{first_token}%"))
        match = cur.fetchone()

        if not match:
            # Project not present — insert it as a hybrid project with CONTD-4 CLEARED.
            cur.execute('SELECT id FROM grid_regions WHERE code = %s', (region,))
            region_id = cur.fetchone()['id']
            plant_code = hybrid_plant_code(comp_rows)
            cur.execute('SELECT id FROM plant_types WHERE code = %s', (plant_code,))
            pt = cur.fetchone()
            if not pt:
                not_found.append(f"[{region}] {station} (no plant_type code={plant_code})")
                continue
            plant_id = pt['id']

            pooling = comp_rows[0].get('pooling')
            pooling_id = None
            if pooling:
                cur.execute('SELECT id FROM pooling_stations WHERE name = %s AND "regionId" = %s', (pooling, region_id))
                ps = cur.fetchone()
                pooling_id = ps['id'] if ps else None

            total_cap = sum((r['comp_total'] or 0) for r in comp_rows)
            project_id = cuid_like()

            if DRY:
                print(f"  ➕ {region} {station[:50]:50}: INSERT new hybrid ({plant_code}, total={total_cap:.1f}) + {len(comp_rows)} phase(s)")
                inserted += 1
                continue

            cur.execute("""
                INSERT INTO generation_projects (id, name, "regionId", "plantTypeId",
                  "poolingStationId", "totalCapacityMw", "createdById", "createdAt", "updatedAt")
                VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
            """, (project_id, station, region_id, plant_id, pooling_id, total_cap, admin_user))
            # CONTD-4 application — CLEARED (since these are FTC-pipeline projects)
            cur.execute("""
                INSERT INTO contd4_applications (id, "projectId", "applicationDate",
                  "capacityApr26Mw", status, "createdAt", "updatedAt")
                VALUES (%s, %s, NOW(), %s, 'CLEARED', NOW(), NOW())
            """, (cuid_like(), project_id, total_cap))
            # Mock match dict so the rest of the flow proceeds
            match = {'id': project_id, 'name': station}
            inserted += 1

        project_id = match['id']
        # Component capacities for the project header
        comp_caps = {'WIND': 0.0, 'SOLAR': 0.0, 'BESS': 0.0, 'PSP': 0.0}
        for r in comp_rows:
            comp_caps[r['source']] = (comp_caps.get(r['source']) or 0) + (r['comp_total'] or 0)

        if DRY:
            print(f"  ✓ {region} {match['name'][:50]:50}: would set "
                  f"solar={comp_caps['SOLAR']:.1f} wind={comp_caps['WIND']:.1f} bess={comp_caps['BESS']:.1f} "
                  f"+ {len(comp_rows)} phase(s)")
            updated += 1
            continue

        # Update project header
        cur.execute("""
            UPDATE generation_projects SET
              "solarCapacityMw" = %s,
              "windCapacityMw"  = %s,
              "bessCapacityMw"  = %s,
              "updatedAt" = NOW()
            WHERE id = %s
        """, (
            comp_caps['SOLAR'] or None,
            comp_caps['WIND']  or None,
            comp_caps['BESS']  or None,
            project_id,
        ))

        # Replace phases with one per source component
        cur.execute('DELETE FROM commissioning_phases WHERE "projectId" = %s', (project_id,))
        for r in comp_rows:
            cur.execute("""
                INSERT INTO commissioning_phases
                  (id, "projectId", "sourceType", "capacityAppliedMw",
                   "ftcCompletedMw", "ftcCompletedDate", "capacityUnderFtcMw",
                   "tocIssuedMw", "tocIssuedDate", "capacityUnderTocMw",
                   "codDeclaredMw", "codDeclaredDate", "capacityPendingCodMw",
                   "expectedApr26Mw", "createdAt", "updatedAt")
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
            """, (
                cuid_like(), project_id, r['source'],
                r['applied'] or 0,
                r['ftc'],     r['ftc_date'],    r['under_ftc'],
                r['toc'],     r['toc_date'],    r['under_toc'],
                r['cod'],     r['cod_date'],    r['cod_pending'],
                r['expected'],
            ))
        print(f"  ✓ {region} {match['name'][:50]:50}: replaced phases ({len(comp_rows)})")
        updated += 1

    print(f"\n→ {updated} hybrid projects synced, {inserted} inserted, {len(not_found)} skipped")
    if not_found:
        print("  Unmatched:")
        for n in not_found[:10]:
            print(f"    - {n}")

    if DRY:
        conn.rollback()
        print("\n[DRY RUN] no changes committed")
    else:
        conn.commit()
        print("\n✅ Committed.")
    conn.close()


if __name__ == "__main__":
    main()
