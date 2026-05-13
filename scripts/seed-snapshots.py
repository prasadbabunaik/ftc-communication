#!/usr/bin/env python3
"""
Seed GridSnapshot records from all available Excel files.
Uses the same extraction + computation logic as validate-all-files.py.
Calls the POST /api/grid/snapshots endpoint for each date.

Usage:
  python3 scripts/seed-snapshots.py
  (server must be running on localhost:3000)
"""

import sys, json, urllib.request, urllib.error
from pathlib import Path
from datetime import datetime

# ── Import the extraction + computation logic from validate-all-files ─────────
sys.path.insert(0, str(Path(__file__).parent))

import importlib.util, types

# We re-implement a trimmed version here to avoid import issues

import openpyxl, re

EXCEL_DIR = Path('public/data/excel')

FILES = [
    ('2026-04-23', 'CONTD and FTC details 23.04.26.xlsx'),
    ('2026-04-24', 'CONTD and FTC details 24.04.26.xlsx'),
    ('2026-04-27', 'CONTD and FTC details 27.04.2026.xlsx'),
    ('2026-04-28', 'CONTD and FTC details 280426.xlsx'),
    ('2026-04-29', 'CONTD and FTC details 290426.xlsx'),
    ('2026-04-30', 'CONTD and FTC details 30.04.xlsx'),
    ('2026-05-02', 'CONTD and FTC details 02.05.26.xlsx'),
    ('2026-05-04', 'CONTD and FTC details 04.05.26.xlsx'),
    ('2026-05-06', 'CONTD and FTC details 06052026.xlsx'),
    ('2026-05-07', 'CONTD and FTC details 07.05.26.xlsx'),
    ('2026-05-11', 'CONTD and FTC details 110526.xlsx'),
    ('2026-05-12', 'CONTD and FTC details 120526.xlsx'),
    ('2026-05-13', 'CONTD and FTC details 130526.xlsx'),
]

REGION_ORDER = ['NR', 'WR', 'SR', 'ER', 'NER']
SOURCE_ORDER = ['WIND', 'SOLAR', 'BESS', 'HYBRID', 'COAL', 'HYDRO', 'PSP']

# ── helpers ───────────────────────────────────────────────────────────────────

def n(v): return float(v) if v else 0.0

def safe_dec(v):
    if v is None: return 0.0
    s = str(v).strip()
    if s in ('#VALUE!','#N/A','#NAME?','#REF!','#DIV/0!','','-','NA','None','0.0'): return 0.0
    try: return float(s)
    except: return 0.0

def safe_str(v):
    if v is None: return None
    s = str(v).strip()
    return s or None

def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s or s in ('-','0','None','0.0','NA','N/A','--'): return None
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
    if m: return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.match(r'(\d{1,2})[-./](\d{1,2})[-./](\d{4})', s)
    if m:
        d,mo,y=int(m.group(1)),int(m.group(2)),int(m.group(3))
        try:
            datetime(y,mo,d)
            return f"{y}-{mo:02d}-{d:02d}"
        except: return None
    return None

def parse_cap_month(v):
    if v is None: return None
    s = str(v).strip()
    MONTHS={'jan':'01','feb':'02','mar':'03','apr':'04','may':'05','jun':'06',
            'jul':'07','aug':'08','sep':'09','oct':'10','nov':'11','dec':'12'}
    m = re.match(r"([A-Za-z]+)['\s-]?(\d{2,4})", s)
    if m:
        mon_str=m.group(1)[:3].lower()
        yr=m.group(2)
        if len(yr)==2: yr='20'+yr
        if mon_str in MONTHS: return f"{yr}-{MONTHS[mon_str]}"
    return None

def is_sr_num(v):
    if v is None: return False
    try: return int(float(str(v))) > 0
    except: return False

def pt_code(label):
    if not label: return 'SOLAR'
    l=str(label).upper()
    if 'WIND' in l and 'SOLAR' in l and 'BESS' in l: return 'HYBRID_WSB'
    if 'WIND' in l and 'SOLAR' in l and 'PSP'  in l: return 'HYBRID_WP'
    if 'SOLAR' in l and 'PSP'  in l:                  return 'HYBRID_SP'
    if 'SOLAR' in l and 'BESS' in l:                  return 'HYBRID_SB'
    if 'WIND'  in l and 'SOLAR' in l:                 return 'HYBRID_WS'
    if 'HYDRO' in l and 'PSP'  in l:                  return 'HYBRID_HP'
    if 'WIND'  in l and 'PSP'  in l:                  return 'HYBRID_WP'
    if 'SOLAR'  in l:                                  return 'SOLAR'
    if 'WIND'   in l:                                  return 'WIND'
    if 'BESS'   in l or 'BATTERY' in l:               return 'BESS'
    if 'PSP'    in l or 'PUMPED'  in l:               return 'PSP'
    if 'HYDRO'  in l:                                  return 'HYDRO'
    if 'COAL'   in l or 'THERMAL' in l:               return 'COAL'
    return 'SOLAR'

def src_type(s):
    if not s: return 'SOLAR'
    u=str(s).upper().strip()
    if 'PSP' in u or 'PUMP' in u: return 'PSP'
    if 'HYDRO' in u: return 'HYDRO'
    if 'COAL' in u or 'THERMAL' in u: return 'COAL'
    if 'BESS' in u or 'BATTERY' in u: return 'BESS'
    if 'SOLAR' in u: return 'SOLAR'
    if 'WIND' in u: return 'WIND'
    return 'SOLAR'

def get_source(p):
    pt=p.get('plantTypeCode','')
    if 'HYBRID' in pt: return 'HYBRID'
    phs=p.get('phases',[])
    if phs: return phs[0].get('sourceType','SOLAR')
    return pt if pt in SOURCE_ORDER else 'SOLAR'

def find_header(ws, kws, start=1, maxr=300, excl=None):
    if isinstance(kws,str): kws=[kws]
    for i,row in enumerate(ws.iter_rows(min_row=start,max_row=maxr,values_only=True),start):
        txt=' '.join(str(v or '') for v in row).lower()
        if all(k.lower() in txt for k in kws):
            if excl and any(e.lower() in txt for e in excl): continue
            return i
    return None

def find_all_headers(ws, kws, start=1, maxr=300, excl=None):
    if isinstance(kws,str): kws=[kws]
    hits=[]
    for i,row in enumerate(ws.iter_rows(min_row=start,max_row=maxr,values_only=True),start):
        txt=' '.join(str(v or '') for v in row).lower()
        if all(k.lower() in txt for k in kws):
            if excl and any(e.lower() in txt for e in excl): continue
            hits.append(i)
    return hits

# ── Extract FTC projects from a sheet ────────────────────────────────────────

def extract_ftc(ws, region_code, title_row, end_row=None):
    data_start=None
    for i in range(title_row+1,title_row+6):
        row=list(ws.iter_rows(min_row=i,max_row=i,values_only=True))[0]
        if is_sr_num(row[0]): data_start=i; break
    if data_start is None: return []
    limit=end_row if end_row else title_row+150
    projects=[]
    for i in range(data_start,limit):
        row=list(ws.iter_rows(min_row=i,max_row=i,values_only=True))[0]
        if not is_sr_num(row[0]): continue
        name=safe_str(row[1])
        if not name: continue
        projects.append({
            'name':name,'region':region_code,
            'plantTypeCode':pt_code(safe_str(row[3])),
            'totalCapacityMw':safe_dec(row[5]),
            'phases':[{
                'sourceType':src_type(safe_str(row[8])),
                'capacityAppliedMw':safe_dec(row[7]),
                'ftcCompletedMw':safe_dec(row[9]),
                'ftcCompletedDate':parse_date(row[10]),
                'proposedFtcDate':parse_date(row[15]),
                'capacityUnderFtcMw':safe_dec(row[16]),
                'tocIssuedMw':safe_dec(row[11]),
                'tocIssuedDate':parse_date(row[12]),
                'capacityUnderTocMw':safe_dec(row[17]),
                'codDeclaredMw':safe_dec(row[13]),
                'codDeclaredDate':parse_date(row[14]),
                'expectedApr26Mw':safe_dec(row[19]),
            }]
        })
    return projects

def extract_contd4(ws, region_code, title_row, end_row, col_offset=0):
    data_start=None
    for i in range(title_row+1, min(title_row+5, end_row)):
        row=list(ws.iter_rows(min_row=i,max_row=i,values_only=True))[0]
        if is_sr_num(row[0]): data_start=i; break
    if data_start is None: return []
    records=[]
    for i in range(data_start, end_row):
        row=list(ws.iter_rows(min_row=i,max_row=i,values_only=True))[0]
        o=col_offset
        if is_sr_num(row[0]):
            name=safe_str(row[1+o])
            cap=safe_dec(row[5+o])
            cap_apr=safe_dec(row[8+o])
            cap_month=parse_cap_month(safe_str(row[7+o]))
            if name and cap>0:
                records.append({
                    'name':name,'region':region_code,
                    'plantTypeCode':pt_code(safe_str(row[4+o])),
                    'totalCapacityMw':cap,
                    'capacityApr26Mw':cap_apr,
                    'capacityMonth':cap_month,
                })
    return records

def extract_tx(ws, region_code, title_row, end_row=None):
    data_start=None
    for i in range(title_row+1,title_row+5):
        row=list(ws.iter_rows(min_row=i,max_row=i,values_only=True))[0]
        if is_sr_num(row[0]): data_start=i; break
    if data_start is None: return []
    limit=end_row if end_row else title_row+100
    elements=[]
    for i in range(data_start,limit):
        row=list(ws.iter_rows(min_row=i,max_row=i,values_only=True))[0]
        if not is_sr_num(row[0]): continue
        ename=safe_str(row[2])
        if not ename: continue
        etype_raw=safe_str(row[3])
        def tx_type(t):
            if not t: return 'LINE'
            u=str(t).upper().strip()
            if u in ('ICT','PT'): return 'ICT'
            if u in ('GT','BR','BAYS','REACTORS'): return 'GT'
            if u=='ST': return 'ST'
            return 'LINE'
        re_flag=safe_str(row[4])
        cap_mva=safe_dec(row[6])
        line_km_raw=row[7]
        def clean_len(v):
            if v is None: return None
            m=re.search(r'(\d+(?:\.\d+)?)',str(v))
            return float(m.group(1)) if m else None
        pending=safe_str(row[9])
        cap_apr=safe_dec(row[11])
        len_apr=clean_len(row[12])
        pending_bool=str(pending or '').strip().upper() in ('YES','Y')
        is_re=str(re_flag or '').strip().upper()=='RE'
        elements.append({
            'region':region_code,'elementName':ename,
            'elementType':tx_type(etype_raw),'isRe':is_re,
            'capacityMva':cap_mva,'lineLengthKm':clean_len(line_km_raw),
            'pendingFtc':pending_bool,'capacityApr26Mva':cap_apr,
            'lineLengthApr26Km':len_apr,
        })
    return elements

def extract_workbook(wb):
    contd4,ftc,tx=[],[],[]
    regions={'NR':0,'WR':1,'ER':0,'NER':0,'SR':0}
    for sheet_name,c4_off in regions.items():
        if sheet_name not in wb.sheetnames: continue
        ws=wb[sheet_name]
        c4_title=find_header(ws,'CONTD-4',start=1,maxr=10)
        ftc_titles=find_all_headers(ws,['Generation Capacity','FTC'],excl=['Transmission','Source wise'],start=1,maxr=200)
        src_wise=find_header(ws,'Source wise Segregation',start=1,maxr=300)
        tx_title=find_header(ws,['Transmission Elements','FTC'],start=1,maxr=300)
        if c4_title:
            ends=[t for t in [(ftc_titles[0] if ftc_titles else None),tx_title] if t and t>c4_title]
            c4_end=min(ends) if ends else c4_title+100
            contd4.extend(extract_contd4(ws,sheet_name,c4_title,c4_end,c4_off))
        if ftc_titles:
            main_ftc=ftc_titles[0]
            cands=[t for t in [tx_title,src_wise] if t and t>main_ftc]
            if len(ftc_titles)>=2: cands.append(ftc_titles[1])
            ftc_end=min(cands) if cands else None
            ftc.extend(extract_ftc(ws,sheet_name,main_ftc,ftc_end))
        if tx_title:
            tx_end=None
            if sheet_name=='SR' and ftc_titles:
                for t in ftc_titles:
                    if t>tx_title+5: tx_end=t; break
            elif src_wise and src_wise>tx_title: tx_end=src_wise
            tx.extend(extract_tx(ws,sheet_name,tx_title,tx_end))
    return contd4,ftc,tx

# ── Compute summary tables ────────────────────────────────────────────────────

def compute_t1(contd4_projects):
    t1={}
    for p in contd4_projects:
        region=p['region']
        source=get_source(p)
        key=f"{region}|{source}"
        if key not in t1: t1[key]={'region':region,'source':source,'totalMw':0.0,'months':{}}
        t1[key]['totalMw']+=p['totalCapacityMw']
        month=p.get('capacityMonth')
        mw=p.get('capacityApr26Mw',0.0)
        if month and mw:
            t1[key]['months'][month]=t1[key]['months'].get(month,0.0)+mw
    all_months=sorted({m for v in t1.values() for m in v['months']})
    rows=[]
    for region in REGION_ORDER:
        region_rows=[t1[f"{region}|{src}"] for src in SOURCE_ORDER if f"{region}|{src}" in t1]
        if not region_rows: continue
        rows.extend(region_rows)
        sub={'region':region,'source':'Total','totalMw':0.0,'months':{},'isSubtotal':True}
        for r in region_rows:
            sub['totalMw']+=r['totalMw']
            for m in all_months: sub['months'][m]=sub['months'].get(m,0.0)+(r['months'].get(m,0.0))
        rows.append(sub)
    all_india={'region':'All India','source':'All Sources','totalMw':0.0,'months':{},'isTotal':True}
    for r in [x for x in rows if x.get('isSubtotal')]:
        all_india['totalMw']+=r['totalMw']
        for m in all_months: all_india['months'][m]=all_india['months'].get(m,0.0)+r['months'].get(m,0.0)
    rows.append(all_india)
    return {'rows':rows,'allMonths':all_months}

def compute_t2(ftc_projects):
    t2={}
    for p in ftc_projects:
        region=p['region']
        source=get_source(p)
        key=f"{region}|{source}"
        if key not in t2:
            t2[key]={'region':region,'source':source,'totalCapacityMw':0,'contd4CapacityMw':0,
                     'appliedMw':0,'ftcApprovedMw':0,'ftcPendingMw':0,
                     'tocIssuedMw':0,'tocPendingMw':0,'codCompletedMw':0,'codPendingMw':0,'expectedMw':0}
        row=t2[key]
        row['totalCapacityMw']+=p['totalCapacityMw']
        for ph in p.get('phases',[]):
            row['appliedMw']+=n(ph.get('capacityAppliedMw'))
            row['ftcApprovedMw']+=n(ph.get('ftcCompletedMw'))
            row['ftcPendingMw']+=n(ph.get('capacityUnderFtcMw'))
            row['tocIssuedMw']+=n(ph.get('tocIssuedMw'))
            row['tocPendingMw']+=n(ph.get('capacityUnderTocMw'))
            cod=n(ph.get('codDeclaredMw'))
            toc=n(ph.get('tocIssuedMw'))
            row['codCompletedMw']+=cod
            row['codPendingMw']+=max(0,toc-cod)
            row['expectedMw']+=n(ph.get('expectedApr26Mw'))
    return t2

def compute_t3(tx_elements):
    t3={}
    for el in tx_elements:
        region=el['region']
        is_re=el['isRe']
        etype=el['elementType']
        if etype=='LINE': cat='LINE_RE' if is_re else 'LINE_NONRE'
        elif etype=='ICT': cat='ICT_RE' if is_re else 'ICT_NONRE'
        else: continue
        key=f"{region}|{cat}"
        if key not in t3: t3[key]={'completedNo':0,'completedKm':0.0,'completedMva':0.0,'pendingNo':0,'pendingKm':0.0,'pendingMva':0.0}
        row=t3[key]
        is_line=(etype=='LINE')
        if not el['pendingFtc']:
            row['completedNo']+=1
            if is_line: row['completedKm']+=(el['lineLengthKm'] or 0.0)
            else: row['completedMva']+=(el['capacityMva'] or 0.0)
        else:
            row['pendingNo']+=1
            if is_line:
                row['pendingKm']+=((el['lineLengthApr26Km'] or 0.0) or (el['lineLengthKm'] or 0.0))
            else:
                row['pendingMva']+=((el['capacityApr26Mva'] or 0.0) or (el['capacityMva'] or 0.0))
    return t3

# ── Upload to server ──────────────────────────────────────────────────────────

def upload_snapshot(date_str, t1, t2, t3, label=None):
    payload = json.dumps({
        'snapshotDate': date_str,
        'label': label or date_str,
        't1Json': t1, 't2Json': t2, 't3Json': t3,
    }).encode('utf-8')
    # Use the API endpoint directly via HTTP
    req = urllib.request.Request(
        'http://localhost:3000/api/grid/snapshots-direct',
        data=payload,
        headers={'Content-Type': 'application/json'},
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read())
    except Exception as e:
        return {'error': str(e)}

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    results = []
    for date_str, filename in FILES:
        fpath = EXCEL_DIR / filename
        if not fpath.exists():
            print(f"  ⚠ {date_str}: file not found — {filename}")
            continue
        try:
            wb = openpyxl.load_workbook(str(fpath), data_only=True)
            contd4, ftc, tx = extract_workbook(wb)
            t1 = compute_t1(contd4)
            t2 = compute_t2(ftc)
            t3 = compute_t3(tx)
            results.append({'date': date_str, 'label': date_str, 't1': t1, 't2': t2, 't3': t3})
            print(f"  ✓ {date_str}: {len(contd4)} CONTD-4, {len(ftc)} FTC, {len(tx)} TX extracted")
        except Exception as e:
            print(f"  ✗ {date_str}: error — {e}")

    # Write to a JSON file for the seeding endpoint to consume
    out_path = Path('scripts/snapshots-seed.json')
    with open(out_path, 'w') as f:
        json.dump(results, f)
    print(f"\nWrote {len(results)} snapshots to {out_path}")
    print("Run: node scripts/seed-snapshots-db.js  to load into DB")

if __name__ == '__main__':
    main()
