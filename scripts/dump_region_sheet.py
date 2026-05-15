"""Dump structure of a per-region sheet (NR) to understand layout for per-project data."""
from pathlib import Path
from openpyxl import load_workbook

DATA_DIR = Path(__file__).parent.parent / "public/data/excel"
FILE = "CONTD and FTC details 130526.xlsx"

wb = load_workbook(DATA_DIR / FILE, data_only=True)

# Inspect NR sheet for Section 2 structure
for sheet_name in ['NR', 'WR', 'SR', 'ER', 'NER']:
    ws = wb[sheet_name]
    print(f"\n{'='*100}")
    print(f"  SHEET: {sheet_name} (max_row={ws.max_row}, max_col={ws.max_column})")
    print(f"{'='*100}\n")
    for r in range(1, min(ws.max_row + 1, 80)):
        row_cells = []
        for c in range(1, min(ws.max_column + 1, 22)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                s = repr(v)[:30]
                row_cells.append(f"C{c}={s}")
        if row_cells:
            print(f"R{r:3}: " + " | ".join(row_cells))
