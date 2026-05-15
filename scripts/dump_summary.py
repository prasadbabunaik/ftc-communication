"""Dump first 250 rows of Summary sheet to understand layout."""
from pathlib import Path
from openpyxl import load_workbook

DATA_DIR = Path(__file__).parent.parent / "public/data/excel"
FILE = "CONTD and FTC details 130526.xlsx"

wb = load_workbook(DATA_DIR / FILE, data_only=True)
print("Sheets:", wb.sheetnames)

for name in wb.sheetnames:
    if "summary" in name.lower():
        ws = wb[name]
        break

print(f"\n=== Summary sheet (max_row={ws.max_row}, max_col={ws.max_column}) ===\n")

for r in range(1, 260):
    row_data = []
    for c in range(1, 20):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            row_data.append(f"C{c}={repr(v)[:40]}")
    if row_data:
        print(f"Row {r:3}: " + " | ".join(row_data))
