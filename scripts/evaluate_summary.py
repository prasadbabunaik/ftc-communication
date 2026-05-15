"""
Extracts Table 2 (FTC Pipeline by Region x Source) from the Summary sheet
of each Excel file and prints a structured comparison.

Run: python3 scripts/evaluate_summary.py
"""

import os, re
from pathlib import Path
import openpyxl
from openpyxl import load_workbook

DATA_DIR = Path(__file__).parent.parent / "public/data/excel"

FILES = {
    "11-May": "CONTD and FTC details 110526.xlsx",
    "12-May": "CONTD and FTC details 120526.xlsx",
    "13-May": "CONTD and FTC details 130526.xlsx",
}

# ── helpers ──────────────────────────────────────────────────────────────────

def cell_val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return str(v) if v else 0.0

def cell_str(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""

# ── Summary sheet Table 2 scanner ─────────────────────────────────────────────
# Table 2 is "Total Generation Capacity Details Under FTC".
# It starts after a merged header row containing that phrase.
# Rows follow pattern: blank|region | source | TotalCap | CONTD4 | Applied |
#   FTCApproved | FTCPending | TOCIssued | TOCPending | CODDone | CODPending | Expected

COLS = {
    # Column index (1-based) will be detected dynamically,
    # but from previous manual inspection the layout is:
    # A=SrNo/blank, B=Region, C=Source, D=TotalCap, E=CONTD4Cap,
    # F=Applied, G=FTCApproved, H=FTCPending, I=TOCIssued, J=TOCPending,
    # K=CODDone, L=CODPending, M=Expected(May/Jun)
    "region":      2,
    "source":      3,
    "total_cap":   4,
    "contd4":      5,
    "applied":     6,
    "ftc_approved":7,
    "ftc_pending": 8,
    "toc_issued":  9,
    "toc_pending": 10,
    "cod_done":    11,
    "cod_pending": 12,
    "expected":    13,
}

KNOWN_REGIONS = {"NR", "WR", "SR", "ER", "NER"}
KNOWN_SOURCES = {"WIND", "SOLAR", "BESS", "HYBRID", "COAL", "HYDRO", "PSP"}

def find_table2_start(ws):
    """Scan for the row containing 'Total Generation Capacity Details Under FTC'."""
    for row in ws.iter_rows(min_row=1, max_row=200):
        for cell in row:
            if cell.value and "Total Generation Capacity" in str(cell.value) and "FTC" in str(cell.value):
                return cell.row
    return None

def parse_table2(ws, start_row):
    """
    Parse Table 2 starting from start_row.
    Returns list of dicts: {region, source, total_cap, contd4, applied,
                             ftc_approved, ftc_pending, toc_issued, toc_pending,
                             cod_done, cod_pending, expected, is_subtotal, is_total}
    """
    rows = []
    # Skip the header rows (usually 2-3 rows after start)
    # Scan until we hit the next major section or 200 rows
    for r in range(start_row + 1, start_row + 200):
        region = cell_str(ws, r, COLS["region"])
        source = cell_str(ws, r, COLS["source"])

        # Stop if we hit the next table header or blank stretch
        if any(kw in region for kw in ["Transmission", "Hybrid", "Source wise", "Monthly"]):
            break
        if any(kw in source for kw in ["Transmission", "Hybrid", "Source wise", "Monthly"]):
            break

        # Detect All India total row
        is_total = "All India" in region or "All India" in source or \
                   "Grand" in region or "Grand" in source or \
                   "GRAND" in region or "GRAND" in source

        # Detect subtotal (region total) rows
        is_subtotal = (region.upper() in KNOWN_REGIONS and source.upper() == "TOTAL") or \
                      (source.upper() == "TOTAL") or \
                      ("Total" in source and region.upper() in KNOWN_REGIONS)

        # Skip empty rows
        applied = cell_val(ws, r, COLS["applied"])
        total_cap = cell_val(ws, r, COLS["total_cap"])
        if not region and not source and total_cap == 0 and applied == 0:
            continue

        rows.append({
            "region":       region,
            "source":       source,
            "total_cap":    total_cap,
            "contd4":       cell_val(ws, r, COLS["contd4"]),
            "applied":      applied,
            "ftc_approved": cell_val(ws, r, COLS["ftc_approved"]),
            "ftc_pending":  cell_val(ws, r, COLS["ftc_pending"]),
            "toc_issued":   cell_val(ws, r, COLS["toc_issued"]),
            "toc_pending":  cell_val(ws, r, COLS["toc_pending"]),
            "cod_done":     cell_val(ws, r, COLS["cod_done"]),
            "cod_pending":  cell_val(ws, r, COLS["cod_pending"]),
            "expected":     cell_val(ws, r, COLS["expected"]),
            "is_subtotal":  is_subtotal,
            "is_total":     is_total,
        })

        if is_total:
            break

    return rows


def extract_all_india_row(rows):
    """Find the All India grand total row."""
    for r in rows:
        if r["is_total"]:
            return r
    return None


def extract_region_totals(rows):
    """Find each region's subtotal row."""
    totals = {}
    for r in rows:
        if r["is_subtotal"] and r["region"].upper() in KNOWN_REGIONS:
            totals[r["region"].upper()] = r
    return totals


def fmt(v):
    if isinstance(v, (int, float)):
        return f"{v:>10.2f}"
    return f"{str(v):>10}"


# ── main ─────────────────────────────────────────────────────────────────────

def analyse_file(label, filename):
    path = DATA_DIR / filename
    if not path.exists():
        print(f"  ❌ File not found: {filename}")
        return

    wb = load_workbook(path, data_only=True)

    # Try "Summary" sheet
    summary_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            summary_sheet = wb[name]
            break
    if summary_sheet is None:
        print(f"  ❌ No Summary sheet in {filename}. Sheets: {wb.sheetnames}")
        return

    ws = summary_sheet

    start = find_table2_start(ws)
    if start is None:
        print(f"  ❌ Could not find Table 2 header in Summary sheet of {filename}")
        # Debug: print first 50 rows col B
        for r in range(1, 50):
            v = cell_str(ws, r, 2)
            if v:
                print(f"    row {r} colB: {v!r}")
        return

    print(f"\n{'='*80}")
    print(f"  {label} — {filename}  (Table 2 header at row {start})")
    print(f"{'='*80}")

    rows = parse_table2(ws, start)

    if not rows:
        print("  ⚠️  No data rows found")
        return

    # Print All India row
    ai = extract_all_india_row(rows)
    if ai:
        print(f"\n  ALL INDIA GRAND TOTAL:")
        print(f"    Total Cap   : {fmt(ai['total_cap'])}")
        print(f"    CONTD-4     : {fmt(ai['contd4'])}")
        print(f"    Applied     : {fmt(ai['applied'])}")
        print(f"    FTC Approved: {fmt(ai['ftc_approved'])}")
        print(f"    FTC Pending : {fmt(ai['ftc_pending'])}")
        print(f"    TOC Issued  : {fmt(ai['toc_issued'])}")
        print(f"    TOC Pending : {fmt(ai['toc_pending'])}")
        print(f"    COD Done    : {fmt(ai['cod_done'])}")
        print(f"    COD Pending : {fmt(ai['cod_pending'])}")
        print(f"    Expected    : {fmt(ai['expected'])}")
    else:
        print("  ⚠️  All India row not found")

    # Print per-region totals
    rt = extract_region_totals(rows)
    if rt:
        print(f"\n  REGION TOTALS (Applied | FTC Approved | TOC Issued | COD Done | Expected):")
        hdr = f"  {'Region':6}  {'Applied':>10}  {'FTC Apprvd':>10}  {'TOC Issued':>10}  {'COD Done':>10}  {'Expected':>10}"
        print(hdr)
        print("  " + "-"*60)
        for rgn in ["NR","WR","SR","ER","NER"]:
            r = rt.get(rgn)
            if r:
                print(f"  {rgn:6}  {fmt(r['applied'])}  {fmt(r['ftc_approved'])}  {fmt(r['toc_issued'])}  {fmt(r['cod_done'])}  {fmt(r['expected'])}")
    else:
        print("  ⚠️  No region subtotals found — printing all rows:")
        for r in rows[:30]:
            print(f"    {r['region']:8} {r['source']:10} applied={r['applied']:>8.2f} ftc={r['ftc_approved']:>8.2f} toc={r['toc_issued']:>8.2f} cod={r['cod_done']:>8.2f}")


def main():
    print("\n" + "="*80)
    print("  FTC PORTAL — Excel Summary Extractor (Table 2: FTC Pipeline by Region×Source)")
    print("="*80)

    for label, filename in FILES.items():
        analyse_file(label, filename)

    print("\n")


if __name__ == "__main__":
    main()
