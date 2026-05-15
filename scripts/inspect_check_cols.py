"""Dump the Summary sheet showing all columns including the 'check' validation columns."""
from pathlib import Path
from openpyxl import load_workbook

DATA = Path(__file__).parent.parent / "public/data/excel"
wb = load_workbook(DATA / "CONTD and FTC details 130526.xlsx", data_only=False)  # keep formulas
ws = wb["Summary"]

print(f"Summary sheet: max_row={ws.max_row}, max_col={ws.max_column}\n")

# Inspect first 6 rows for headers
print("=== HEADER ROWS ===")
for r in range(1, 8):
    for c in range(1, min(ws.max_column + 1, 25)):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            print(f"  R{r:2} C{c:2}: {repr(v)[:60]}")
    print()

# Now dump rows 4-35 with all columns including formulas
print("\n=== TABLE 1: CONTD-4 STUDY (rows 4-35) — all cols incl. formulas ===")
for r in range(4, 36):
    cells = []
    for c in range(1, min(ws.max_column + 1, 14)):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            cells.append(f"C{c}={repr(v)[:35]}")
    if cells:
        print(f"R{r:2}: " + " | ".join(cells))

print("\n=== TABLE 2: FTC PIPELINE (rows 36-90) — all cols incl. formulas ===")
for r in range(36, 90):
    cells = []
    for c in range(1, min(ws.max_column + 1, 16)):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            cells.append(f"C{c}={repr(v)[:40]}")
    if cells:
        print(f"R{r:2}: " + " | ".join(cells))
