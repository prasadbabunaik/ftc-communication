"""
Parse Section 2 (FTC Pipeline) from per-region sheets of the May 13 Excel,
extract per-project values, and produce a mapping that can be used to update DB.

Section 2 layout:
  Header is at row 22 (NR) or row 42 (WR/SR/ER/NER)
  Data rows follow until next section header (~ row 50 onwards)

Per-project columns:
  C2: Generating Station name
  C3: Pooling Station
  C4: Plant Type
  C5: Region
  C6: Total Plant Capacity
  C7: Total Capacity (CONTD4 issued)
  C8: Capacity applied for FTC
  C9: Sources Type Applied for FTC
  C10: FTC Completed Capacity (MW)
  C11: FTC date if completed
  C12: TOC Issued Capacity (MW)
  C13: TOC issuance date
  C14: COD declared Capacity (MW)
  C15: COD Date
  C17: Capacity Under Process for FTC
  C18: Capacity Under Process for TOC
  C19: Capacity Pending for COD
  C20: Capacity expected to commission (in target month)
"""

import json
from pathlib import Path
from openpyxl import load_workbook

DATA_DIR = Path(__file__).parent.parent / "public/data/excel"
FILE = "CONTD and FTC details 130526.xlsx"

# Header row for Section 2 per region (1-indexed)
SECTION2_HEADER_ROW = {
    "NR":  22,   # Header at R22, data R23+
    "WR":  42,   # Header at R42, data R43+
    "SR":  None, # Will detect dynamically
    "ER":  None,
    "NER": None,
}

# Maximum scan distance from header
MAX_ROWS = 50

def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s in ("-", "NA", "N/A") or "not applied" in s.lower():
        return None
    try:
        return float(s)
    except ValueError:
        return None

def text(v):
    if v is None:
        return None
    return str(v).strip() if str(v).strip() else None

def find_section2_header(ws):
    """Find row where Section 2 ('Generation Capacity Under Process of FTC') header starts."""
    for r in range(1, 80):
        for c in range(1, 6):
            v = ws.cell(row=r, column=c).value
            if v and "Generation Capacity Under Process of FTC" in str(v):
                return r + 1  # data starts after header row
            if v and "Plant Type" in str(v) and "Wind" in str(v):
                return r  # column header row — data starts next row
    return None

def parse_section2(ws, region_code):
    """Parse Section 2 rows for a region. Returns list of project dicts."""
    # Find Plant Type header
    header_row = None
    for r in range(1, 80):
        v = ws.cell(row=r, column=4).value
        if v and "Plant Type" in str(v):
            header_row = r
            break
    if header_row is None:
        print(f"  ⚠️  Could not find Section 2 header in {region_code}")
        return []

    projects = []
    for r in range(header_row + 1, header_row + MAX_ROWS):
        name = text(ws.cell(row=r, column=2).value)
        if not name:
            continue
        # Stop when we hit next section (transmission elements)
        if "Transmission" in name or "Hybrid Capacity" in name or "Sr. No" in name:
            break

        plant_type = text(ws.cell(row=r, column=4).value)
        if not plant_type:
            continue

        proj = {
            "row":             r,
            "region":          region_code,
            "name":            name,
            "pooling_station": text(ws.cell(row=r, column=3).value),
            "plant_type":      plant_type,
            "total_cap":       num(ws.cell(row=r, column=6).value),
            "contd4_cap":      num(ws.cell(row=r, column=7).value),
            "applied":         num(ws.cell(row=r, column=8).value),
            "source_types":    text(ws.cell(row=r, column=9).value),
            "ftc_completed":   num(ws.cell(row=r, column=10).value),
            "ftc_date_raw":    ws.cell(row=r, column=11).value,
            "toc_issued":      num(ws.cell(row=r, column=12).value),
            "toc_date_raw":    ws.cell(row=r, column=13).value,
            "cod_declared":    num(ws.cell(row=r, column=14).value),
            "cod_date_raw":    ws.cell(row=r, column=15).value,
            "under_ftc":       num(ws.cell(row=r, column=17).value),
            "under_toc":       num(ws.cell(row=r, column=18).value),
            "cod_pending":     num(ws.cell(row=r, column=19).value),
            "expected":        num(ws.cell(row=r, column=20).value),
        }
        projects.append(proj)

    return projects

def main():
    wb = load_workbook(DATA_DIR / FILE, data_only=True)
    all_projects = {}

    for region in ["NR", "WR", "SR", "ER", "NER"]:
        ws = wb[region]
        projects = parse_section2(ws, region)
        all_projects[region] = projects
        print(f"\n{region}: {len(projects)} projects")
        for p in projects:
            applied = p["applied"] or 0
            ftc     = p["ftc_completed"] or 0
            toc     = p["toc_issued"] or 0
            cod     = p["cod_declared"] or 0
            exp     = p["expected"] or 0
            uft     = p["under_ftc"] or 0
            print(f"  {p['name'][:50]:50}  app={applied:>7.2f}  ftc={ftc:>7.2f}  toc={toc:>7.2f}  cod={cod:>7.2f}  uft={uft:>5.2f}  exp={exp:>6.2f}")

    # Save to JSON for next step
    # Strip datetime objects
    def clean(p):
        return {k: v for k, v in p.items() if k not in ("ftc_date_raw","toc_date_raw","cod_date_raw","row")}
    json_data = {region: [clean(p) for p in projects] for region, projects in all_projects.items()}
    out = Path(__file__).parent / "may13_projects.json"
    out.write_text(json.dumps(json_data, indent=2, default=str))
    print(f"\n✓ Saved per-project data to {out}")


if __name__ == "__main__":
    main()
