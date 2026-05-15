"""
Parse Section 1 (CONTD-4 study) from per-region sheets of the May 13 Excel
and update each project's contd4.capacityMonth + capacityApr26Mw to the
Excel-stated values, so the dashboard's CONTD-4 study table shows the
correct distribution of expected completion months.

Section 1 columns (typical):
  C2: Generating Station name
  C5/C6: Generation Type (NR uses C5)
  C6/C7: Capacity(MW)         — NR uses C6
  C7/C8: Application Date     — NR uses C7
  C8/C9: Proposed FTC date    — NR uses C8 (e.g. "April'26", "May'26", "June'26", datetime)
  C9/C10: Capacity(MW) to be completed in [Month]'26 — NR uses C9
"""

import sys
import re
from pathlib import Path
from datetime import datetime, date
import psycopg2, psycopg2.extras
from openpyxl import load_workbook

DATA_DIR = Path(__file__).parent.parent / "public/data/excel"
FILE     = "CONTD and FTC details 130526.xlsx"
DB_URL   = "postgresql://postgres:S0perg%4026@10.5.133.55:5432/ftc_communication"
DRY_RUN  = "--dry-run" in sys.argv

# Column maps per region (Section 1 layout differs slightly).
# WR has an extra "Name of Developer" column which shifts everything by one.
COLS = {
    "NR":  {"name": 2, "type": 5, "cap": 6, "proposed": 8, "month_cap": 9},
    "WR":  {"name": 3, "type": 6, "cap": 7, "proposed": 9, "month_cap": 10},
    "SR":  {"name": 2, "type": 5, "cap": 6, "proposed": 8, "month_cap": 9},
    "ER":  {"name": 2, "type": 5, "cap": 6, "proposed": 8, "month_cap": 9},
    "NER": {"name": 2, "type": 5, "cap": 6, "proposed": 8, "month_cap": 9},
}

# Section 1 column header row (containing 'Generating Station' / 'Capacity(MW)').
# In all sheets this is row 3 — data starts at row 4 and runs until the next
# section ("Generation Capacity Under Process of FTC") begins.
HEADER_ROW = 3

MONTH_MAP = {
    'jan': 1, 'january': 1, 'feb': 2, 'february': 2,
    'mar': 3, 'march': 3,   'apr': 4, 'april': 4,
    'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7,
    'aug': 8, 'august': 8, 'sep': 9, 'sept': 9, 'september': 9,
    'oct': 10, 'october': 10, 'nov': 11, 'november': 11,
    'dec': 12, 'december': 12,
}

def parse_proposed_month(v):
    """Convert a 'Proposed FTC date' cell to 'YYYY-MM' or None."""
    if v is None: return None
    if isinstance(v, (datetime, date)):
        return f"{v.year}-{v.month:02d}"
    s = str(v).strip()
    if not s or s in ('-', 'NA', 'N/A'):
        return None

    # Match patterns like "April'26", "May 26", "June'2026", "May'26"
    m = re.search(r"([A-Za-z]+)['\s\-]?(\d{2,4})", s)
    if m:
        month_str = m.group(1).lower()
        year_str  = m.group(2)
        month_num = MONTH_MAP.get(month_str)
        if month_num:
            year = int(year_str)
            if year < 100: year += 2000
            return f"{year}-{month_num:02d}"

    # "DD-MM-YYYY" → just take MM
    m = re.match(r"\d{1,2}[-/](\d{1,2})[-/](\d{2,4})", s)
    if m:
        month_num = int(m.group(1))
        year = int(m.group(2))
        if year < 100: year += 2000
        return f"{year}-{month_num:02d}"

    return None

def parse_num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    try: return float(s)
    except ValueError: return None

def parse_section1(ws, region):
    cols = COLS[region]
    projects = []

    for r in range(HEADER_ROW + 1, HEADER_ROW + 40):
        name = ws.cell(row=r, column=cols["name"]).value
        # Stop when we hit Section 2 banner
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c).value
            if cell and isinstance(cell, str) and 'Process of FTC' in cell:
                return projects

        if not name: continue
        name = str(name).strip()
        if not name:
            continue

        cap_total    = parse_num(ws.cell(row=r, column=cols["cap"]).value)
        proposed_raw = ws.cell(row=r, column=cols["proposed"]).value
        proposed_ym  = parse_proposed_month(proposed_raw)
        month_cap    = parse_num(ws.cell(row=r, column=cols["month_cap"]).value)

        if cap_total is None and month_cap is None and not proposed_ym:
            continue   # skip blank lines

        projects.append({
            "row": r,
            "region": region,
            "name": name,
            "total_capacity": cap_total,
            "proposed_month": proposed_ym,
            "month_capacity": month_cap or 0.0,
        })
    return projects

def main():
    print(f"\n{'='*78}\n  Sync CONTD-4 months {'(DRY RUN)' if DRY_RUN else ''}\n{'='*78}")
    wb = load_workbook(DATA_DIR / FILE, data_only=True)

    all_projects = []
    for region in ["NR", "WR", "SR", "ER", "NER"]:
        ws = wb[region]
        rows = parse_section1(ws, region)
        all_projects.extend(rows)
        print(f"\n{region}: {len(rows)} CONTD-4 projects in Section 1")
        for p in rows:
            print(f"  - {p['name'][:50]:50}  total={p['total_capacity'] or 0:>7.1f}  "
                  f"month={p['proposed_month'] or '—':>7}  cap_in_month={p['month_capacity']:>6.1f}")

    print(f"\n→ {len(all_projects)} total projects parsed")

    # Connect to DB and update
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    cur  = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    updates = 0
    not_found = []

    for p in all_projects:
        if not p["proposed_month"]:
            continue   # skip projects with no target month

        # Find DB project — fuzzy match: first 30 chars of name + region
        first_part = p["name"][:30]
        cur.execute("""
            SELECT gp.id AS project_id, gp.name, ca.id AS contd4_id, ca.status
            FROM generation_projects gp
            JOIN grid_regions gr ON gp."regionId" = gr.id
            JOIN contd4_applications ca ON ca."projectId" = gp.id
            WHERE gp.name ILIKE %s AND gr.code = %s
              AND ca.status NOT IN ('CLEARED', 'REJECTED')
        """, (f"%{first_part}%", p["region"]))
        rows = cur.fetchall()
        if not rows:
            not_found.append(f"{p['region']} | {p['name']}")
            continue

        contd4_id = rows[0]["contd4_id"]
        new_month = p["proposed_month"]
        new_cap   = p["month_capacity"]

        if DRY_RUN:
            print(f"  ✓ {p['region']} {rows[0]['name'][:45]:45} → month={new_month}, cap={new_cap}")
        else:
            cur.execute("""
                UPDATE contd4_applications
                SET "capacityMonth" = %s,
                    "capacityApr26Mw" = %s,
                    "updatedAt" = NOW()
                WHERE id = %s
            """, (new_month, new_cap, contd4_id))
        updates += 1

    print(f"\n→ {updates} projects updated")
    if not_found:
        print(f"\n⚠️  {len(not_found)} projects not found in DB:")
        for n in not_found[:10]:
            print(f"    - {n}")
        if len(not_found) > 10:
            print(f"    ...and {len(not_found)-10} more")

    if DRY_RUN:
        conn.rollback()
        print("\n[DRY RUN] no changes committed")
    else:
        conn.commit()
        print("\n✅ Changes committed.")
    conn.close()


if __name__ == "__main__":
    main()
