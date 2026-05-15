"""
Full evaluation: parse Excel Summary Table 2 (Region × Source FTC Pipeline)
AND query the database, then run our computation logic and compare.

The key question: does our computePipelineMatrix / buildPipelineRows logic
produce the same numbers as the Excel Summary sheet for a given snapshot?

Since the DB holds one latest state (not a per-date snapshot),
we compare DB-computed values against the LATEST Excel (13-May).

Run: python3 scripts/evaluate_vs_excel.py
"""

import os, sys
from pathlib import Path
import psycopg2
import psycopg2.extras
from openpyxl import load_workbook
from datetime import datetime

DATA_DIR = Path(__file__).parent.parent / "public/data/excel"
DB_URL   = "postgresql://postgres:S0perg%4026@10.5.133.55:5432/ftc_communication"

# ── Excel parsing ─────────────────────────────────────────────────────────────

def num(v):
    if v is None or (isinstance(v, str) and '#' in v):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0

def parse_table2_from_excel(filepath):
    """
    Parse Table 2 (FTC Pipeline by Region × Source) from Summary sheet.
    Returns {(region, source): {col: value, ...}}

    Column mapping from dump:
      C1=Region, C2=Source, C3=TotalCap, C4=CONTD4, C5=Applied,
      C6=FTCApproved, C7=FTCPending, C8=TOCIssued, C9=TOCPending,
      C10=CODDone, C11=CODPending, C12=Expected
    Table starts at row 36 (header) → data rows 38..85
    """
    wb = load_workbook(filepath, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        return {}

    # Find "Total Generation Capacity Details" header row
    table2_start = None
    for r in range(1, 200):
        v = ws.cell(row=r, column=1).value
        if v and "Total Generation Capacity" in str(v) and "FTC" in str(v):
            table2_start = r
            break
    if table2_start is None:
        return {}

    data = {}
    cur_region = None
    for r in range(table2_start + 2, table2_start + 80):
        c1 = ws.cell(row=r, column=1).value
        c2 = ws.cell(row=r, column=2).value
        if c1:
            c1s = str(c1).strip()
            if c1s in ("NR", "WR", "SR", "ER", "NER"):
                cur_region = c1s
            elif c1s == "All India":
                cur_region = "All India"

        c2s = str(c2).strip() if c2 else ""
        if not c2s:
            continue

        row = {
            "region":       cur_region,
            "source":       c2s,
            "total_cap":    num(ws.cell(row=r, column=3).value),
            "contd4":       num(ws.cell(row=r, column=4).value),
            "applied":      num(ws.cell(row=r, column=5).value),
            "ftc_approved": num(ws.cell(row=r, column=6).value),
            "ftc_pending":  num(ws.cell(row=r, column=7).value),
            "toc_issued":   num(ws.cell(row=r, column=8).value),
            "toc_pending":  num(ws.cell(row=r, column=9).value),
            "cod_done":     num(ws.cell(row=r, column=10).value),
            "cod_pending":  num(ws.cell(row=r, column=11).value),
            "expected":     num(ws.cell(row=r, column=12).value),
        }

        if cur_region and c2s:
            data[(cur_region, c2s)] = row

        if cur_region == "All India" and c2s == "Total":
            break

    return data

# Normalize source names: Excel uses "Wind","Solar","BESS","Hybrid","Coal","Hydro","PSP"
# Our DB uses "WIND","SOLAR","BESS","HYBRID","COAL","HYDRO","PSP"
SOURCE_MAP = {
    "Wind":  "WIND",
    "Solar": "SOLAR",
    "BESS":  "BESS",
    "Hybrid":"HYBRID",
    "Coal":  "COAL",
    "Hydro": "HYDRO",
    "PSP":   "PSP",
}

# ── DB queries ────────────────────────────────────────────────────────────────

def query_db():
    conn = psycopg2.connect(DB_URL)
    cur  = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Fetch all commissioning phases with project + region + plant type
    cur.execute("""
        SELECT
            gr.code AS region,
            UPPER(pt.label) AS plant_type,
            gp."totalCapacityMw"  AS total_capacity_mw,
            gp.id AS project_id,
            ca."capacityApr26Mw"  AS contd4_cap,
            ca.status AS contd4_status,
            cp."capacityAppliedMw"   AS capacity_applied_mw,
            cp."sourceType"          AS source_type,
            cp."ftcCompletedMw"      AS ftc_completed_mw,
            cp."ftcCompletedDate"    AS ftc_completed_date,
            cp."capacityUnderFtcMw"  AS capacity_under_ftc_mw,
            cp."tocIssuedMw"         AS toc_issued_mw,
            cp."tocIssuedDate"       AS toc_issued_date,
            cp."capacityUnderTocMw"  AS capacity_under_toc_mw,
            cp."codDeclaredMw"       AS cod_declared_mw,
            cp."codDeclaredDate"     AS cod_declared_date,
            cp."capacityPendingCodMw" AS capacity_pending_cod_mw,
            cp."expectedApr26Mw"     AS expected_apr26_mw
        FROM generation_projects gp
        JOIN grid_regions gr ON gp."regionId" = gr.id
        JOIN plant_types pt ON gp."plantTypeId" = pt.id
        LEFT JOIN contd4_applications ca ON ca."projectId" = gp.id
        LEFT JOIN commissioning_phases cp ON cp."projectId" = gp.id
        WHERE ca.status = 'CLEARED'
        ORDER BY gr.code, pt.label, gp.id, cp."sourceType"
    """)
    phases = cur.fetchall()

    # Also fetch total capacity per project (for projects with CLEARED CONTD4)
    cur.execute("""
        SELECT
            gr.code AS region,
            UPPER(pt.label) AS plant_type,
            gp."totalCapacityMw"  AS total_capacity_mw,
            ca."capacityApr26Mw"  AS contd4_cap
        FROM generation_projects gp
        JOIN grid_regions gr ON gp."regionId" = gr.id
        JOIN plant_types pt ON gp."plantTypeId" = pt.id
        LEFT JOIN contd4_applications ca ON ca."projectId" = gp.id
        WHERE ca.status = 'CLEARED'
    """)
    projects = cur.fetchall()

    conn.close()
    return phases, projects


def normalize_source(source_type, plant_type):
    """
    Map DB phase → pipeline bucket, matching JS getProjectSource() logic exactly:
      - Hybrid projects (isHybrid=true) → always 'HYBRID' regardless of phase sourceType
      - Otherwise → use phase sourceType directly
    plant_type here is UPPER(pt.label), so we check for 'HYBRID' substring.
    """
    pt = plant_type.upper() if plant_type else ""
    if "HYBRID" in pt:
        return "HYBRID"
    if source_type:
        return source_type.upper()
    # Fallback from label
    if "WIND" in pt:
        return "WIND"
    if "SOLAR" in pt:
        return "SOLAR"
    if "BESS" in pt or "BATTERY" in pt:
        return "BESS"
    if "COAL" in pt or "THERMAL" in pt:
        return "COAL"
    if "HYDRO" in pt:
        return "HYDRO"
    if "PSP" in pt or "PUMP" in pt:
        return "PSP"
    return pt

REGIONS     = ["NR", "WR", "SR", "ER", "NER"]
SOURCES_DB  = ["WIND", "SOLAR", "BESS", "HYBRID", "COAL", "HYDRO", "PSP"]

def compute_from_db(phases, projects):
    """
    Re-implement computePipelineMatrix logic in Python.
    Returns {(region, source): {applied, ftc_approved, ftc_pending, toc_issued, toc_pending, cod_done, cod_pending, expected, total_cap, contd4}}
    """
    # Step 1: aggregate totalCapacityMw and contd4Cap per (region, source)
    project_cap = {}  # (region, source) → {total_cap, contd4}
    for p in projects:
        region   = p["region"]
        source   = normalize_source(None, p["plant_type"])
        key      = (region, source)
        if key not in project_cap:
            project_cap[key] = {"total_cap": 0.0, "contd4": 0.0}
        project_cap[key]["total_cap"] += float(p["total_capacity_mw"] or 0)
        project_cap[key]["contd4"]    += float(p["contd4_cap"] or 0)

    # Step 2: aggregate phase data per (region, source)
    phase_agg = {}
    for ph in phases:
        region = ph["region"]
        source = normalize_source(ph["source_type"], ph["plant_type"])
        key    = (region, source)
        if key not in phase_agg:
            phase_agg[key] = {
                "applied":      0.0,
                "ftc_approved": 0.0,
                "ftc_pending":  0.0,
                "toc_issued":   0.0,
                "toc_pending":  0.0,
                "cod_done":     0.0,
                "cod_pending":  0.0,
                "expected":     0.0,
            }
        a = phase_agg[key]
        a["applied"]      += float(ph["capacity_applied_mw"]  or 0)
        a["ftc_approved"] += float(ph["ftc_completed_mw"]     or 0)
        a["ftc_pending"]  += float(ph["capacity_under_ftc_mw"] or 0)
        a["toc_issued"]   += float(ph["toc_issued_mw"]        or 0)
        a["toc_pending"]  += float(ph["capacity_under_toc_mw"] or 0)
        a["cod_done"]     += float(ph["cod_declared_mw"]      or 0)
        a["expected"]     += float(ph["expected_apr26_mw"]    or 0)

        toc = float(ph["toc_issued_mw"] or 0)
        cod = float(ph["cod_declared_mw"] or 0)
        stored_cod_pending = ph["capacity_pending_cod_mw"]
        if stored_cod_pending is not None:
            a["cod_pending"] += float(stored_cod_pending)
        else:
            a["cod_pending"] += max(0.0, toc - cod)

    # Step 3: merge
    result = {}
    all_keys = set(list(project_cap.keys()) + list(phase_agg.keys()))
    for key in all_keys:
        caps = project_cap.get(key, {"total_cap": 0.0, "contd4": 0.0})
        agg  = phase_agg.get(key, {
            "applied":0, "ftc_approved":0, "ftc_pending":0,
            "toc_issued":0, "toc_pending":0, "cod_done":0,
            "cod_pending":0, "expected":0
        })
        result[key] = {**caps, **agg}

    return result


# ── Comparison ────────────────────────────────────────────────────────────────

COLS_TO_CHECK = [
    ("applied",      "C5 Applied"),
    ("ftc_approved", "C6 FTC Approved"),
    ("ftc_pending",  "C7 FTC Pending"),
    ("toc_issued",   "C8 TOC Issued"),
    ("toc_pending",  "C9 TOC Pending"),
    ("cod_done",     "C10 COD Done"),
    ("cod_pending",  "C11 COD Pending"),
    ("expected",     "C12 Expected"),
]

TAINTED_CELLS = {
    # NR TOC values are formula errors in the Excel (circular refs/REF errors → huge numbers)
    # We skip comparison for those cells
    ("NR", "Solar", "toc_issued"),
    ("NR", "BESS",  "toc_issued"),
    ("NR", "Hybrid","toc_issued"),
    ("NR", "Coal",  "toc_issued"),
    ("NR", "PSP",   "toc_issued"),
    ("NR", "Wind",  "toc_issued"),
    ("NR", "Hydro", "toc_issued"),
    ("NR", "Total", "toc_issued"),
}

def compare(excel_data, db_data, label="13-May"):
    print(f"\n{'='*90}")
    print(f"  COMPARISON: {label} Excel Summary vs DB-computed values")
    print(f"  (Only CLEARED projects' phases are included)")
    print(f"{'='*90}")

    any_mismatch = False

    for region in REGIONS:
        region_mismatches = []
        for src_excel, src_db in SOURCE_MAP.items():
            excel_row = excel_data.get((region, src_excel))
            db_row    = db_data.get((region, src_db))

            if excel_row is None and db_row is None:
                continue  # both zero — skip

            for col, col_label in COLS_TO_CHECK:
                excel_val = excel_row[col] if excel_row else 0.0
                db_val    = db_row.get(col, 0.0) if db_row else 0.0

                # Skip known tainted cells
                if (region, src_excel, col) in TAINTED_CELLS:
                    continue

                diff = abs(excel_val - db_val)
                if diff > 0.05:  # tolerance
                    region_mismatches.append({
                        "region": region,
                        "source": src_db,
                        "col": col_label,
                        "excel": excel_val,
                        "db":    db_val,
                        "diff":  diff,
                    })
                    any_mismatch = True

        if region_mismatches:
            print(f"\n  ❌ {region} — {len(region_mismatches)} mismatch(es):")
            for m in region_mismatches:
                print(f"     {m['source']:8} {m['col']:20} Excel={m['excel']:>10.2f}  DB={m['db']:>10.2f}  Δ={m['diff']:>8.2f}")
        else:
            # Check if region has any data at all
            has_data = any(
                excel_data.get((region, s)) or db_data.get((region, sd))
                for s, sd in SOURCE_MAP.items()
            )
            if has_data:
                print(f"\n  ✅ {region} — all values match")
            else:
                print(f"\n  ⬜ {region} — no data in either source")

    # All India totals
    print(f"\n  --- ALL INDIA TOTALS ---")
    ai_excel = excel_data.get(("All India", "Total")) or {}
    ai_db    = {col: sum(db_data.get((r, s), {}).get(col, 0.0) for r in REGIONS for s in SOURCES_DB)
                for col, _ in COLS_TO_CHECK}

    for col, label_ in COLS_TO_CHECK:
        if "toc_issued" in col:
            print(f"    {'All India':10} {label_:20} Excel={ai_excel.get(col,0):>10.2f}  DB={ai_db.get(col,0):>10.2f}  (TOC may have Excel formula errors in NR rows)")
            continue
        ev = ai_excel.get(col, 0.0)
        dv = ai_db.get(col, 0.0)
        diff = abs(ev - dv)
        sym  = "✅" if diff <= 0.05 else "❌"
        print(f"    {sym} All India   {label_:20} Excel={ev:>12.2f}  DB={dv:>12.2f}  Δ={diff:>8.2f}")

    if not any_mismatch:
        print("\n  🎉 No mismatches found (excluding known NR TOC formula errors)")
    else:
        print(f"\n  ⚠️  Mismatches found — check above")


def print_db_summary(db_data):
    print(f"\n{'='*90}")
    print(f"  DATABASE SUMMARY (from current DB state)")
    print(f"  Format: Region | Source | Applied | FTC Apprvd | TOC Issued | COD Done | Expected")
    print(f"{'='*90}")
    for region in REGIONS:
        has_any = False
        for src in SOURCES_DB:
            r = db_data.get((region, src))
            if r and (r.get("applied", 0) + r.get("total_cap", 0)) > 0:
                if not has_any:
                    print(f"\n  {region}:")
                    has_any = True
                print(f"    {src:8}  applied={r['applied']:>9.2f}  ftc={r['ftc_approved']:>9.2f}  "
                      f"toc={r['toc_issued']:>9.2f}  cod={r['cod_done']:>9.2f}  exp={r['expected']:>9.2f}")
        if not has_any:
            print(f"\n  {region}: (no data)")

    # Totals
    print(f"\n  ALL INDIA:")
    for col, clabel in COLS_TO_CHECK:
        total = sum(db_data.get((r, s), {}).get(col, 0.0) for r in REGIONS for s in SOURCES_DB)
        print(f"    {clabel:22}: {total:>12.2f}")


def print_excel_summary(excel_data, label):
    print(f"\n{'='*90}")
    print(f"  EXCEL SUMMARY ({label}) — Region Totals from Summary Sheet")
    print(f"{'='*90}")
    for region in REGIONS:
        total_row = excel_data.get((region, "Total"))
        if total_row and total_row["applied"] > 0:
            print(f"  {region}: Applied={total_row['applied']:>9.2f}  FTC={total_row['ftc_approved']:>9.2f}  "
                  f"TOC={total_row['toc_issued']:>9.2f}  COD={total_row['cod_done']:>9.2f}  Exp={total_row['expected']:>9.2f}")
        else:
            print(f"  {region}: (no data or zero)")
    ai = excel_data.get(("All India", "Total"))
    if ai:
        print(f"  ALL INDIA: Applied={ai['applied']:>9.2f}  FTC={ai['ftc_approved']:>9.2f}  "
              f"TOC={ai['toc_issued']:>9.2f}  COD={ai['cod_done']:>9.2f}  Exp={ai['expected']:>9.2f}")


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    print("\nFTC Portal — Logic Evaluation")
    print(f"  Comparing: DB state vs Excel Summary sheet (May 13 as primary reference)")
    print()

    # Parse all 3 Excel files
    files = {
        "11-May": "CONTD and FTC details 110526.xlsx",
        "12-May": "CONTD and FTC details 120526.xlsx",
        "13-May": "CONTD and FTC details 130526.xlsx",
    }

    excel_datasets = {}
    for label, fname in files.items():
        path = DATA_DIR / fname
        print(f"  Parsing Excel {label}... ", end="", flush=True)
        data = parse_table2_from_excel(path)
        excel_datasets[label] = data
        print(f"  {len(data)} region×source cells found")

    # Query DB
    print("\n  Querying database... ", end="", flush=True)
    try:
        phases, projects = query_db()
        print(f"  {len(phases)} phases, {len(projects)} projects (CLEARED)")
    except Exception as e:
        print(f"  ❌ DB connection failed: {e}")
        # Still print Excel summaries
        for label, data in excel_datasets.items():
            print_excel_summary(data, label)
        return

    db_data = compute_from_db(phases, projects)

    # Print DB summary
    print_db_summary(db_data)

    # Print each Excel's region totals
    for label, data in excel_datasets.items():
        print_excel_summary(data, label)

    # Compare against the most recent file (13-May)
    compare(excel_datasets["13-May"], db_data, "13-May")

    # Also check if values changed between 11/12/13-May
    print(f"\n{'='*90}")
    print("  CHANGE DETECTION: Did Summary values change between May 11-12-13?")
    print(f"{'='*90}")
    cols_check = [("applied","Applied"), ("ftc_approved","FTC Approved"),
                  ("toc_issued","TOC Issued"), ("cod_done","COD Done")]
    for region in REGIONS:
        for src_excel, src_db in SOURCE_MAP.items():
            vals = []
            for label in ["11-May","12-May","13-May"]:
                r = excel_datasets[label].get((region, src_excel))
                if r:
                    vals.append(tuple(r.get(c, 0) for c, _ in cols_check))
            if not vals or all(v == (0,0,0,0) for v in vals):
                continue
            if len(set(vals)) > 1:  # changed across days
                print(f"\n  {region} {src_db} changed:")
                for label, val in zip(["11-May","12-May","13-May"], vals):
                    print(f"    {label}: " + "  ".join(f"{l}={v:.2f}" for (_, l), v in zip(cols_check, val)))


if __name__ == "__main__":
    main()
